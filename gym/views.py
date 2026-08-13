from datetime import timedelta, timezone, datetime
import json
from plistlib import InvalidFileException
import re
from django.db import models
from django.db.models import Q, Count, Sum as DjangoSum
from django.http import JsonResponse
from django.urls import reverse
from django.utils import timezone
from decimal import Decimal, InvalidOperation
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction,  IntegrityError
from django.contrib.auth.hashers import make_password
import openpyxl
from sympy import Sum
from django.core.paginator import Paginator
from django.utils.dateparse import parse_date
from django.conf import settings 
from .models import (
    Almacen, DetallePack, IngresoMonetario, Kardex, MetodoPago, Modulo, MovimientoCaja, PagoVenta, PermisoRol, ProductoVariante, Rol, PlanEmpresa, Empresa, Sucursal, TipoProducto, Usuario, CanalVenta, UnidadMedida, Category,
    Producto, PrecioProducto, Stock, TipoIngreso, Ingreso, DetalleIngreso, Turno,
    Caja, CajaTurno, TipoEgreso, Egreso, DetalleEgreso, Proveedor, Compra, DetalleCompra,
    Venta, DetalleVenta, Traspaso, DetalleTraspaso, EgresoMonetario, Plan, Cliente,
    Pago, Asistencia, Membresia,
)
from django.core.exceptions import ValidationError
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.http import StreamingHttpResponse
import time
from gym.decorators import permiso_requerido

# ====================================================
#  ROL
# ====================================================
@login_required
@permiso_requerido('rol_list', 'ver')
def rol_list(request):
    roles = Rol.objects.filter(estado=True, fk_empresa=request.user.sucursal.fk_empresa).order_by('-fecha_creacion')
    return render(request, 'usuarios/rol_list.html', {'roles': roles})


@login_required
@permiso_requerido('rol_list', 'crear')
def rol_create(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()

        if not nombre:
            messages.error(request, 'El nombre del rol es obligatorio.')
            return redirect('rol_list')

        if Rol.objects.filter(nombre__iexact=nombre, estado=True, fk_empresa=request.user.sucursal.fk_empresa).exists():
            messages.error(request, f'Ya existe un rol activo con el nombre "{nombre}".')
            return redirect('rol_list')

        Rol.objects.create(nombre=nombre, estado=True, fk_empresa=request.user.sucursal.fk_empresa)
        messages.success(request, 'Rol creado correctamente.')
    else:
        messages.error(request, 'Método no permitido.')
        
    return redirect('rol_list')

@login_required
@permiso_requerido('rol_list', 'ver')
def rol_permisos(request, rol_id):
    rol = get_object_or_404(Rol, pk=rol_id, estado=True)
    
    # 1. Traemos todos los activos
    todos_los_modulos = Modulo.objects.filter(is_active=True).order_by('orden')
    
    # 2. Lógica para ordenar: Padre seguido de sus Hijos
    modulos_ordenados = []
    padres = [m for m in todos_los_modulos if m.modulo_padre is None]
    
    for padre in padres:
        modulos_ordenados.append(padre) # Metemos al padre
        hijos = [h for h in todos_los_modulos if h.modulo_padre_id == padre.id]
        modulos_ordenados.extend(hijos) # Metemos a sus hijos justo debajo

    if request.method == 'POST':
        # Usamos los modulos ordenados para el guardado también
        for modulo in modulos_ordenados:
            permiso, _ = PermisoRol.objects.get_or_create(rol=rol, modulo=modulo)
            permiso.puede_ver      = f'ver_{modulo.id}'      in request.POST
            permiso.puede_crear    = f'crear_{modulo.id}'    in request.POST
            permiso.puede_editar   = f'editar_{modulo.id}'   in request.POST
            permiso.puede_eliminar = f'eliminar_{modulo.id}' in request.POST
            permiso.save()

        messages.success(request, f'Permisos del rol "{rol.nombre}" actualizados.')
        return redirect('rol_list')

    permisos_dict = {p.modulo_id: p for p in PermisoRol.objects.filter(rol=rol)}

    modulos_con_permisos = []
    for m in modulos_ordenados: # Usamos la lista ya mezclada
        p = permisos_dict.get(m.id)
        modulos_con_permisos.append({
            'modulo':         m,
            'puede_ver':      p.puede_ver      if p else False,
            'puede_crear':    p.puede_crear    if p else False,
            'puede_editar':   p.puede_editar   if p else False,
            'puede_eliminar': p.puede_eliminar if p else False,
        })

    return render(request, 'usuarios/rol_permisos.html', {
        'rol':                  rol,
        'modulos_con_permisos': modulos_con_permisos,
    })

@login_required
@permiso_requerido('rol_list', 'editar')
def rol_edit(request):

    if request.method == 'POST':
        id = request.POST.get('id')
        rol = get_object_or_404(Rol, pk=id)
        nombre = request.POST.get('nombre', '').strip()

        if not nombre:
            messages.error(request, 'El nombre del rol no puede quedar vacío.')
            return redirect('rol_list')

        if Rol.objects.filter(nombre__iexact=nombre, estado=True).exclude(id=rol.id).exists():
            messages.error(request, f'Ya existe otro rol activo con el nombre "{nombre}".')
            return redirect('rol_list')

        rol.nombre = nombre
        rol.save()
        messages.success(request, 'Rol actualizado correctamente.')
    else:
        messages.error(request, 'Error al procesar la solicitud.')
        
    return redirect('rol_list')

@login_required
@permiso_requerido('rol_list', 'eliminar')
def rol_delete(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        rol = get_object_or_404(Rol, pk=id)
        
        if rol.usuario_set.filter(is_active=True).exists():
             messages.error(request, 'No puedes eliminar un rol que tiene usuarios asignados.')
             return redirect('rol_list')

        rol.estado = False
        rol.save()
        messages.success(request, 'Rol eliminado correctamente.')
        
    return redirect('rol_list')

# ====================================================
#  EMPRESA
# ====================================================
@login_required
@permiso_requerido('empresa_edit', 'ver')
def empresa_list(request):
    empresas = Empresa.objects.select_related('fk_plan_empresa').all().order_by('-fecha_creacion')
    return render(request, 'empresa/list.html', {'empresas': empresas})

@login_required
@permiso_requerido('empresa_edit', 'crear')
def empresa_create(request):
    planes = PlanEmpresa.objects.filter(estado=True)
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        rubro = request.POST.get('rubro')
        fk_plan = request.POST.get('fk_plan_empresa') or None
        Empresa.objects.create(
            nombre=nombre,
            rubro=rubro,
            fk_plan_empresa_id=fk_plan
        )
        messages.success(request, 'Empresa creada correctamente.')
        return redirect('empresa_list')
    return render(request, 'empresa/create.html', {'planes': planes})


def setup_empresa_inicial(empresa, sucursal):
    """
    Se llama UNA vez al crear una empresa.
    Crea todo lo necesario para que pueda operar de inmediato.
    """
    Caja.objects.create(
        nombre='Caja Principal',
        sucursal=sucursal,
        saldo_inicial=0,
        fk_empresa=empresa
    )
    Almacen.objects.create(
        nombre='Almacén Principal',
        sucursal=sucursal
    )
    CanalVenta.objects.bulk_create([
        CanalVenta(nombre='Mostrador', fk_empresa=empresa),
    ])
    UnidadMedida.objects.create(
        nombre='Unidad', abreviatura='und', fk_empresa=empresa
    )
    Category.objects.create(
        name='General', fk_empresa=empresa
    )
    MetodoPago.objects.bulk_create([
        MetodoPago(nombre='Efectivo', empresa=empresa),
        MetodoPago(nombre='QR',       empresa=empresa),
        MetodoPago(nombre='Tarjeta',  empresa=empresa),
    ])
    TipoIngreso.objects.bulk_create([
        TipoIngreso(nombre='Compra',          fk_empresa=empresa),
        TipoIngreso(nombre='Ajuste positivo', fk_empresa=empresa),
    ])
    TipoEgreso.objects.bulk_create([
        TipoEgreso(nombre='Gasto operativo', fk_empresa=empresa),
        TipoEgreso(nombre='Ajuste negativo', fk_empresa=empresa),
    ])
    Turno.objects.create(
        nombre='Turno Mañana',
        hora_inicio='08:00',
        hora_fin='20:00',
        fk_empresa=empresa
    )
    Proveedor.objects.create(
        nombre='Proveedor General',
        contacto='', telefono='', email='', direccion='',
        empresa=empresa
    )
    
def registro_empresa(request):
    # plan gratuito por defecto (14 días)
    plan_prueba = 1
    
    if request.method == 'POST':
        nombre_empresa = request.POST.get('nombre_empresa', '').strip()
        rubro          = request.POST.get('rubro', '').strip()
        moneda         = request.POST.get('moneda', 'BOB')
        simbolo_moneda = request.POST.get('simbolo_moneda', 'Bs.')
        pie_ticket     = request.POST.get('pie_ticket', '').strip()
        nombre         = request.POST.get('nombre', '').strip()
        apellido       = request.POST.get('apellido', '').strip()
        email = request.POST.get('email', '').strip().lower()
        #username       = request.POST.get('username', '').strip()
        password1      = request.POST.get('password1', '')
        password2      = request.POST.get('password2', '')

        errores = []
        if not nombre_empresa: errores.append('El nombre de la empresa es obligatorio.')
        if not rubro:          errores.append('El rubro es obligatorio.')
        if not nombre:         errores.append('Tu nombre es obligatorio.')
        #if not username:       errores.append('El usuario es obligatorio.')
        if not email:          errores.append('El email es obligatorio.')
        if password1 != password2: errores.append('Las contraseñas no coinciden.')
        if len(password1) < 6:     errores.append('Mínimo 6 caracteres en la contraseña.')
        if Usuario.objects.filter(email=email).exists():
            errores.append('Ese email ya está registrado.')

        if errores:
            for e in errores:
                messages.error(request, e)
            return render(request, 'login.html', {
                'plan_prueba': plan_prueba,
            })

        try:
            with transaction.atomic():
                empresa = Empresa.objects.create(
                    nombre=nombre_empresa,
                    rubro=rubro,
                    moneda=moneda,
                    simbolo_moneda=simbolo_moneda,
                    pie_ticket=pie_ticket,
                    fk_plan_empresa=PlanEmpresa.objects.get(id=1),
                    fecha_inicio_plan=timezone.now()
                )
                sucursal = Sucursal.objects.create(
                    nombre='Sucursal Principal',
                    fk_empresa=empresa
                )
                setup_empresa_inicial(empresa, sucursal)

                rol_admin = Rol.objects.create(
                    nombre='Administrador',
                    fk_empresa=empresa
                )
                usuario = Usuario(
                    email=email,
                    nombre=nombre,
                    apellido=apellido,
                    rol=rol_admin,
                    sucursal=sucursal,
                    is_staff=True
                )
                usuario.set_password(password1)
                usuario.save()

        except Exception as ex:
            messages.error(request, f'Error: {ex}')
            return render(request, 'login.html', {'plan_prueba': plan_prueba})

        messages.success(request, f'¡Empresa "{nombre_empresa}" creada! Iniciá sesión.')
        return redirect('login')

    return render(request, 'login.html', {'plan_prueba': plan_prueba})

@login_required
@permiso_requerido('empresa_edit', 'editar')
def empresa_edit(request):
    try:
        empresa = request.user.sucursal.fk_empresa
    except AttributeError:
        messages.error(request, 'No tienes empresa asignada.')
        return redirect('empresa_edit')

    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        if not nombre:
            messages.error(request, 'El nombre de la empresa es obligatorio.')
            return redirect('empresa_edit')

        empresa.nombre = nombre
        empresa.pie_ticket = request.POST.get('pie_ticket', '').strip()
        empresa.moneda = request.POST.get('moneda', empresa.moneda)
        empresa.simbolo_moneda = request.POST.get('simbolo_moneda', empresa.simbolo_moneda).strip() or empresa.simbolo_moneda
        # rubro y plan quedan de solo lectura, no se editan desde aquí
        empresa.save()

        messages.success(request, 'Empresa actualizada correctamente.')
        return redirect('empresa_edit')

    contexto = {
        'empresa': empresa,
        'propietario': empresa.propietario,
    }
    return render(request, 'empresa/empresa_edit.html', contexto)

@login_required
@permiso_requerido('empresa_edit', 'eliminar')
def empresa_delete(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        empresa = get_object_or_404(Empresa, pk=id)
        empresa.estado = False
        empresa.save()
        messages.success(request, 'Empresa desactivada correctamente.')
    return redirect('empresa_list')

# ====================================================
#  SUCURSAL
# ====================================================
@login_required
@permiso_requerido('sucursal_list', 'ver')
def sucursal_list(request):
    empresa = request.user.sucursal.fk_empresa
    sucursales = Sucursal.objects.filter(fk_empresa=empresa, estado=True).order_by('-fecha_creacion')
    return render(request, 'empresa/lista_sucursal.html', {'sucursales': sucursales})

@login_required
@permiso_requerido('sucursal_list', 'crear')
def sucursal_create(request):
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        nombre = request.POST.get('nombre', '').strip()
        direccion = request.POST.get('direccion', '').strip()

        # 1. Validación: Nombre obligatorio
        if not nombre:
            messages.error(request, 'El nombre de la sucursal es obligatorio.')
            return redirect('sucursal_list')

        # 2. Validación: No duplicar nombres en la misma empresa
        if Sucursal.objects.filter(nombre__iexact=nombre, fk_empresa=empresa, estado=True).exists():
            messages.error(request, f'Ya tienes una sucursal activa llamada "{nombre}".')
            return redirect('sucursal_list')

        sucursal = Sucursal.objects.create(
            nombre=nombre,
            direccion=direccion,
            fk_empresa=empresa,
            estado=True
        )
        messages.success(request, 'Sucursal creada correctamente.')
    
    return redirect('sucursal_list')

@login_required
@permiso_requerido('sucursal_list', 'editar')
def sucursal_edit(request):
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        id = request.POST.get('id')
        sucursal = get_object_or_404(Sucursal, pk=id, fk_empresa=empresa)
        
        nombre = request.POST.get('nombre', '').strip()
        direccion = request.POST.get('direccion', '').strip()

        # 1. Validación: Nombre no vacío
        if not nombre:
            messages.error(request, 'El nombre no puede estar vacío.')
            return redirect('sucursal_list')

        # 2. Validación: Nombre duplicado en OTRA sucursal
        if Sucursal.objects.filter(nombre__iexact=nombre, fk_empresa=empresa, estado=True).exclude(id=sucursal.id).exists():
            messages.error(request, f'Otra sucursal ya usa el nombre "{nombre}".')
            return redirect('sucursal_list')

        sucursal.nombre = nombre
        sucursal.direccion = direccion
        sucursal.save()
        messages.success(request, 'Sucursal actualizada correctamente.')
    else:
        messages.error(request, 'Método no permitido.')

    return redirect('sucursal_list')

@login_required 
@permiso_requerido('sucursal_list', 'eliminar')
def sucursal_delete(request):
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        id = request.POST.get('id')
        sucursal = get_object_or_404(Sucursal, pk=id, fk_empresa=empresa)

        if sucursal.id == request.user.sucursal.id:
            messages.error(request, 'No puedes eliminar la sucursal en la que estás trabajando actualmente.')
            return redirect('sucursal_list')

        sucursal.estado = False
        # sucursal._usuario_actual = request.user
        sucursal.save()
        messages.success(request, 'Sucursal eliminada correctamente.')
        
    return redirect('sucursal_list')

# ====================================================
#  ALMACEN
# ====================================================

@login_required
@permiso_requerido('almacen_list', 'ver')
def almacen_list(request):
    # Obtenemos la empresa del usuario actual
    empresa = request.user.sucursal.fk_empresa
    # Filtramos almacenes que pertenecen a sucursales de esa empresa
    almacenes = Almacen.objects.filter(
        sucursal__fk_empresa=empresa, 
        is_active=True
    ).order_by('sucursal', 'nombre')
    
    # También necesitamos las sucursales para el modal de "Nuevo Almacén"
    sucursales = Sucursal.objects.filter(fk_empresa=empresa, estado=True)
    
    return render(request, 'empresa/lista_almacen.html', {
        'almacenes': almacenes,
        'sucursales': sucursales
    })

@login_required
@permiso_requerido('almacen_list', 'crear')
def almacen_create(request):
    empresa = request.user.sucursal.fk_empresa
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        sucursal_id = request.POST.get('sucursal')
        
        # Validamos que la sucursal pertenezca a la empresa
        sucursal = get_object_or_404(Sucursal, id=sucursal_id, fk_empresa=empresa)
        
        # --- VALIDACIÓN EXTRA ---
        if Almacen.objects.filter(nombre__iexact=nombre, sucursal=sucursal, is_active=True).exists():
            messages.error(request, f'Ya existe un almacén llamado "{nombre}" en la sucursal {sucursal.nombre}.')
            return redirect('almacen_list')

        almacen = Almacen.objects.create(
            nombre=nombre,
            descripcion=descripcion,
            sucursal=sucursal,
            is_active=True
        )
        almacen._usuario_actual = request.user
        messages.success(request, 'Almacén creado correctamente.')
    return redirect('almacen_list')

@login_required
@permiso_requerido('almacen_list', 'editar')
def almacen_edit(request):
    empresa = request.user.sucursal.fk_empresa
    if request.method == 'POST':
        id = request.POST.get('id')
        almacen = get_object_or_404(Almacen, pk=id, sucursal__fk_empresa=empresa)
        
        sucursal_id = request.POST.get('sucursal')
        sucursal = get_object_or_404(Sucursal, id=sucursal_id, fk_empresa=empresa)
        nombre = request.POST.get('nombre', '').strip()

        # --- VALIDACIÓN EXTRA (Evitar duplicados al editar) ---
        if Almacen.objects.filter(nombre__iexact=nombre, sucursal=sucursal, is_active=True).exclude(id=almacen.id).exists():
            messages.error(request, f'Ya hay otro almacén con ese nombre en esa sucursal.')
            return redirect('almacen_list')
        
        almacen.nombre = nombre
        almacen.descripcion = request.POST.get('descripcion', '').strip()
        almacen.sucursal = sucursal
        almacen._usuario_actual = request.user
        almacen.save()
        
        messages.success(request, 'Almacén actualizado correctamente.')
    return redirect('almacen_list')

@login_required
@permiso_requerido('almacen_list', 'eliminar')
def almacen_delete(request):
    empresa = request.user.sucursal.fk_empresa
    if request.method == 'POST':
        id = request.POST.get('id')
        almacen = get_object_or_404(Almacen, pk=id, sucursal__fk_empresa=empresa)
        
        # Borrado lógico
        almacen.is_active = False
        almacen._usuario_actual = request.user
        almacen.save()
        messages.success(request, 'Almacén eliminado correctamente.')
    return redirect('almacen_list')

# ====================================================
#  USUARIO
# ====================================================
@login_required
@permiso_requerido('usuario_list', 'ver')
def usuario_list(request):
    empresa = request.user.sucursal.fk_empresa
    usuarios = Usuario.objects.select_related('rol', 'sucursal').filter(sucursal__fk_empresa=empresa, is_active=True) .order_by('-created_at')
 
    roles = Rol.objects.filter(estado=True)
    sucursales = Sucursal.objects.filter(fk_empresa=empresa, estado=True)
 
    return render(request, 'usuarios/lista_usuario.html', {
        'usuarios': usuarios, 'roles': roles, 'sucursales': sucursales,
    })
 
@login_required
@permiso_requerido('usuario_list', 'crear')
@transaction.atomic
def usuario_create(request):
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        email = request.POST.get('email', '').strip().lower()
        password = request.POST.get('password')
        sucursal_id = request.POST.get('sucursal')
        rol_id = request.POST.get('rol')
        nombre = request.POST.get('nombre', '').strip()
 
        # 1. Campos obligatorios
        if not all([email, password, sucursal_id, rol_id, nombre]):
            messages.error(request, 'Nombre, email, contraseña, sucursal y rol son obligatorios.')
            return redirect('usuario_list')
 
        # 2. Email único a nivel GLOBAL (así es como quedó el modelo,
        #    no busques scoped a empresa: el campo ya tiene unique=True)
        if Usuario.objects.filter(email__iexact=email).exists():
            messages.error(request, f'Ya existe una cuenta con el email "{email}".')
            return redirect('usuario_list')
 
        # 3. Seguridad: la sucursal debe ser de la empresa del que está creando
        if not Sucursal.objects.filter(id=sucursal_id, fk_empresa=empresa).exists():
            messages.error(request, 'Sucursal no válida.')
            return redirect('usuario_list')
 
        Usuario.objects.create_user(
            email=email,
            password=password,
            nombre=nombre,
            apellido=request.POST.get('apellido', '').strip(),
            rol_id=rol_id,
            sucursal_id=sucursal_id,
        )
        messages.success(request, 'Usuario creado correctamente.')
 
    return redirect('usuario_list')
 
@login_required
@permiso_requerido('usuario_list', 'editar')
def usuario_edit(request):
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        id = request.POST.get('id')
        usuario = get_object_or_404(Usuario, pk=id, sucursal__fk_empresa=empresa)
 
        email = request.POST.get('email', '').strip().lower()
        sucursal_id = request.POST.get('sucursal')
        new_password = request.POST.get('password')
 
        if not email:
            messages.error(request, 'El email no puede estar vacío.')
            return redirect('usuario_list')
 
        # Email único, excluyendo al propio usuario que se está editando
        if Usuario.objects.filter(email__iexact=email).exclude(id=usuario.id).exists():
            messages.error(request, 'Ese email ya está en uso por otra cuenta.')
            return redirect('usuario_list')
 
        if not Sucursal.objects.filter(id=sucursal_id, fk_empresa=empresa).exists():
            messages.error(request, 'Sucursal no válida.')
            return redirect('usuario_list')
 
        usuario.email = email
        usuario.nombre = request.POST.get('nombre', '').strip()
        usuario.apellido = request.POST.get('apellido', '').strip()
        usuario.rol_id = request.POST.get('rol')
        usuario.sucursal_id = sucursal_id
 
        if new_password:
            usuario.set_password(new_password)
 
        usuario.save()
        messages.success(request, 'Usuario actualizado correctamente.')
 
    return redirect('usuario_list')
 
@login_required
@permiso_requerido('usuario_list', 'eliminar')
def usuario_delete(request):
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        id = request.POST.get('id')
 
        if int(id) == request.user.id:
            messages.error(request, 'No puedes eliminar tu propio usuario.')
            return redirect('usuario_list')
 
        usuario = get_object_or_404(Usuario, pk=id, sucursal__fk_empresa=empresa)
        usuario.is_active = False
        usuario.save()
        messages.success(request, 'Usuario eliminado correctamente.')
 
    return redirect('usuario_list')

# ====================================================
#  CANAL DE VENTA
# ====================================================

@login_required
@permiso_requerido('canalventa_list', 'ver')
def canalventa_list(request):
    canales = CanalVenta.objects.filter(is_active=True, fk_empresa=request.user.sucursal.fk_empresa).order_by('-created_at')
    return render(request, 'ventas/lista_canales.html', {'canales': canales})

@login_required
@permiso_requerido('canalventa_list', 'crear')
@transaction.atomic
def canalventa_create(request):
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion')

        # VALIDACIÓN 1: Nombre vacío
        if not nombre:
            messages.error(request, 'El nombre del canal no puede estar vacío.')
            return redirect('canalventa_list')

        # VALIDACIÓN 2: Duplicados activos para la misma empresa
        if CanalVenta.objects.filter(nombre__iexact=nombre, is_active=True, fk_empresa=empresa).exists():
            messages.error(request, f'Ya existe un canal activo con el nombre "{nombre}".')
            return redirect('canalventa_list')

        # Si pasa las validaciones, se crea
        canal = CanalVenta.objects.create(
            nombre=nombre,
            descripcion=descripcion,
            is_active=True,
            fk_empresa=empresa
        )
        messages.success(request, 'Canal de venta creado correctamente.')
        
    return redirect('canalventa_list')

@login_required
@permiso_requerido('canalventa_list', 'editar')
def canalventa_edit(request):
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        id = request.POST.get('id')
        canal = get_object_or_404(CanalVenta, pk=id, fk_empresa=empresa)
        
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion')

        # VALIDACIÓN 1: Nombre vacío
        if not nombre:
            messages.error(request, 'El nombre no puede quedar vacío.')
            return redirect('canalventa_list')

        # VALIDACIÓN 2: Duplicado (excluyendo el registro actual)
        if CanalVenta.objects.filter(nombre__iexact=nombre, is_active=True, fk_empresa=empresa).exclude(id=canal.id).exists():
            messages.error(request, f'Ya existe otro canal con el nombre "{nombre}".')
            return redirect('canalventa_list')

        canal.nombre = nombre
        canal.descripcion = descripcion
        canal.save()
        messages.success(request, 'Canal de venta actualizado correctamente.')
        
    return redirect('canalventa_list')

@login_required
@permiso_requerido('canalventa_list', 'eliminar')
def canalventa_delete(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        canal = get_object_or_404(CanalVenta, pk=id, fk_empresa=request.user.sucursal.fk_empresa)
        canal.is_active = False  # soft delete
        canal._usuario_actual = request.user
        canal.save()
        messages.success(request, 'Canal de venta eliminado correctamente.')
    return redirect('canalventa_list')

# ====================================================
#  UNIDAD DE MEDIDA
# ====================================================
@login_required
@permiso_requerido('unidademedida_list', 'ver')
def unidadmedida_list(request):
    unidades = UnidadMedida.objects.filter(is_active=True, fk_empresa=request.user.sucursal.fk_empresa).order_by('-created_at')
    return render(request, 'inventario/lista_unidades.html', {
        'unidades': unidades,
    })

@login_required
@permiso_requerido('unidademedida_list', 'crear')
def unidadmedida_create(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre').strip()
        abreviatura = request.POST.get('abreviatura').strip()

        # Validación de duplicado activo
        if UnidadMedida.objects.filter(nombre__iexact=nombre, is_active=True, fk_empresa=request.user.sucursal.fk_empresa).exists():
            messages.error(request, f'Ya existe una unidad de medida activa con el nombre "{nombre}".')
            return redirect('unidadmedida_list')

        UnidadMedida.objects.create(
            nombre=nombre,
            abreviatura=abreviatura,
            is_active=True,
            fk_empresa=request.user.sucursal.fk_empresa
        )

        messages.success(request, 'Unidad de medida creada correctamente.')

    return redirect('unidadmedida_list')

@login_required
@permiso_requerido('unidademedida_list', 'editar')
def unidadmedida_edit(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        unidad = get_object_or_404(UnidadMedida, pk=id, fk_empresa=request.user.sucursal.fk_empresa)

        nombre = request.POST.get('nombre').strip()
        abreviatura = request.POST.get('abreviatura').strip()

        # Validación: duplicado activo excepto sí misma
        if UnidadMedida.objects.filter(nombre__iexact=nombre, is_active=True, fk_empresa=request.user.sucursal.fk_empresa).exclude(id=unidad.id).exists():
            messages.error(request, f'Otra unidad de medida activa ya usa el nombre "{nombre}".')
            return redirect('unidadmedida_list')

        unidad.nombre = nombre
        unidad.abreviatura = abreviatura
        unidad.is_active = True
        unidad._usuario_actual = request.user
        unidad.save()

        messages.success(request, 'Unidad de medida actualizada correctamente.')

    return redirect('unidadmedida_list')

@login_required
@permiso_requerido('unidademedida_list', 'eliminar')
def unidadmedida_delete(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        unidad = get_object_or_404(UnidadMedida, pk=id, fk_empresa=request.user.sucursal.fk_empresa)

        unidad.is_active = False
        unidad._usuario_actual = request.user
        unidad.save()
        messages.success(request, 'Unidad de medida eliminada correctamente.')

    return redirect('unidadmedida_list')

# ====================================================
#  CATEGORÍA
# ====================================================
@login_required
@permiso_requerido('categoria_list', 'ver')
def categoria_list(request):
    categorias = Category.objects.filter(is_active=True, fk_empresa=request.user.sucursal.fk_empresa).order_by('-created_at')
    return render(request, 'inventario/lista_categorias.html', {
        'categorias': categorias
    })

@login_required
@permiso_requerido('categoria_list', 'crear')
def categoria_create(request):
    if request.method == 'POST':
        name = request.POST.get('name').strip()
        description = request.POST.get('description')

        # Validación: evitar duplicados activos
        if Category.objects.filter(name__iexact=name, is_active=True, fk_empresa=request.user.sucursal.fk_empresa).exists():
            messages.error(request, f'Ya existe una categoría activa con el nombre "{name}".')
            return redirect('categoria_list')

        Category.objects.create(
            name=name,
            description=description,
            is_active=True,
            fk_empresa=request.user.sucursal.fk_empresa
        )

        messages.success(request, 'Categoría creada correctamente.')
        return redirect('categoria_list')

    return redirect('categoria_list')

@login_required
@permiso_requerido('categoria_list', 'editar')
def categoria_edit(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        categoria = get_object_or_404(Category, pk=id, fk_empresa=request.user.sucursal.fk_empresa)

        name = request.POST.get('name').strip()
        description = request.POST.get('description')

        # Validación: evitar duplicados activos en otros registros
        if Category.objects.filter(name__iexact=name, is_active=True, fk_empresa=request.user.sucursal.fk_empresa).exclude(id=categoria.id).exists():
            messages.error(request, f'Otra categoría activa ya usa el nombre "{name}".')
            return redirect('categoria_list')

        categoria.name = name
        categoria.description = description
        categoria.is_active = True
        categoria._usuario_actual = request.user
        categoria.save()

        messages.success(request, 'Categoría actualizada correctamente.')
        return redirect('categoria_list')

    return redirect('categoria_list')

@login_required
@permiso_requerido('categoria_list', 'eliminar')
def categoria_delete(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        categoria = get_object_or_404(Category, pk=id, fk_empresa=request.user.sucursal.fk_empresa)
        categoria.is_active = False
        categoria._usuario_actual = request.user
        categoria.save()

        messages.success(request, 'Categoría eliminada correctamente.')

    return redirect('categoria_list')

# ====================================================
#  PRODUCTO INSUMOS
# ====================================================


def safe_decimal(value, default='0'):
    try:
        return Decimal(value) if value not in (None, '') else Decimal(default)
    except InvalidOperation:
        return Decimal(default)

def get_tipo_producto(codigo):
    """Busca el TipoProducto por su código fijo (nunca por nombre o id)."""
    return get_object_or_404(TipoProducto, codigo=codigo, is_active=True)

CODIGO_INSUMO = 'INS-RAW'

@login_required
@permiso_requerido('lista_insumos', 'ver')
def lista_insumos(request):
    insumos = (
        Producto.objects
        .filter(fk_tipo_producto__codigo=CODIGO_INSUMO, is_active=True, fk_empresa=request.user.sucursal.fk_empresa)
        .select_related('unidad_medida', 'category')
        .prefetch_related('variantes')
        .order_by('nombre')
    )

    context = {
        'productos': insumos,
        'unidades_medida': UnidadMedida.objects.filter(is_active=True, fk_empresa=request.user.sucursal.fk_empresa),
        'categorias': Category.objects.filter(is_active=True, fk_empresa=request.user.sucursal.fk_empresa),
        'titulo': 'Insumos',
    }
    return render(request, 'inventario/insumo_list.html', context)

@login_required
@permiso_requerido('lista_insumos', 'crear')
def crear_insumo(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                empresa = request.user.sucursal.fk_empresa
                nombre = request.POST.get('nombre', '').strip()
                sku = request.POST.get('sku', '').strip()
                unidad_medida_id = request.POST.get('unidad_medida')

                if not nombre:
                    raise ValueError("El nombre del insumo es obligatorio.")
                if not sku:
                    raise ValueError("El SKU es obligatorio.")
                if not unidad_medida_id:
                    raise ValueError("La unidad de medida es obligatoria.")
                if ProductoVariante.objects.filter(sku=sku).exists():
                    raise ValueError(f"El SKU '{sku}' ya está en uso.")

                tipo_insumo = get_tipo_producto(CODIGO_INSUMO)

                producto = Producto.objects.create(
                    nombre=nombre,
                    descripcion=request.POST.get('descripcion', ''),
                    fk_empresa=empresa,
                    fk_tipo_producto=tipo_insumo,
                    unidad_medida_id=unidad_medida_id,
                    category_id=request.POST.get('categoria') or None,
                    unidades_por_caja=int(request.POST.get('unidades_por_caja', 0) or 0),
                    tara_por_caja=safe_decimal(request.POST.get('tara_por_caja')),
                )

                # Insumo = siempre variante única
                ProductoVariante.objects.create(
                    producto=producto,
                    nombre_variante="Único",
                    sku=sku,
                    codigo_barras=request.POST.get('codigo_barras') or None,
                    costo=safe_decimal(request.POST.get('costo')),
                    precio_referencial=safe_decimal(request.POST.get('precio_referencial')),
                    maneja_stock=request.POST.get('maneja_stock') == 'on',
                )

                messages.success(request, f'✅ Insumo "{nombre}" creado correctamente.')

        except ValueError as e:
            messages.error(request, f'❌ {e}')
        except IntegrityError:
            messages.error(request, '❌ Ya existe un registro con ese SKU.')
        except Exception as e:
            messages.error(request, f'❌ Error al crear el insumo: {e}')

    return redirect('lista_insumos')


@login_required
@permiso_requerido('lista_insumos', 'editar')
def insumo_edit(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                producto_id = request.POST.get('id')
                variante_id = request.POST.get('variante_id')

                # El filtro por codigo='INS-RAW' evita que esta view
                # pueda editar por error un producto de otro tipo.
                producto = get_object_or_404(
                    Producto, id=producto_id, fk_tipo_producto__codigo=CODIGO_INSUMO, fk_empresa=request.user.sucursal.fk_empresa
                )
                variante = get_object_or_404(
                    ProductoVariante, id=variante_id, producto=producto
                )

                nombre = request.POST.get('nombre', '').strip()
                sku = request.POST.get('sku', '').strip()
                unidad_medida_id = request.POST.get('unidad_medida')

                if not nombre:
                    raise ValueError("El nombre del insumo es obligatorio.")
                if not sku:
                    raise ValueError("El SKU es obligatorio.")
                if not unidad_medida_id:
                    raise ValueError("La unidad de medida es obligatoria.")
                if ProductoVariante.objects.filter(sku=sku).exclude(id=variante.id).exists():
                    raise ValueError(f"El SKU '{sku}' ya está en uso por otro producto.")

                producto.nombre = nombre
                producto.descripcion = request.POST.get('descripcion', '')
                producto.unidad_medida_id = unidad_medida_id
                producto.category_id = request.POST.get('categoria') or None
                producto.unidades_por_caja = int(request.POST.get('unidades_por_caja', 0) or 0)
                producto.tara_por_caja = safe_decimal(request.POST.get('tara_por_caja'))
                producto.save()

                variante.sku = sku
                variante.codigo_barras = request.POST.get('codigo_barras') or None
                variante.costo = safe_decimal(request.POST.get('costo'))
                variante.precio_referencial = safe_decimal(request.POST.get('precio_referencial'))
                variante.maneja_stock = request.POST.get('maneja_stock') == 'on'
                variante.save()

                messages.success(request, f'✅ Insumo "{nombre}" actualizado correctamente.')

        except ValueError as e:
            messages.error(request, f'❌ {e}')
        except IntegrityError:
            messages.error(request, '❌ Ya existe un registro con ese SKU.')
        except Exception as e:
            messages.error(request, f'❌ Error al actualizar el insumo: {e}')

    return redirect('lista_insumos')


@login_required
@permiso_requerido('lista_insumos', 'eliminar')
def insumo_delete(request):
    """
    Baja lógica, no borrado físico.
    Un insumo puede estar referenciado desde DetalleReceta (on_delete=PROTECT)
    o DetallePack, así que borrarlo de verdad puede reventar esas relaciones
    o dejar recetas/combos rotos. Por eso solo se desactiva.
    """
    if request.method == 'POST':
        try:
            producto_id = request.POST.get('id')
            producto = get_object_or_404(
                Producto, id=producto_id, fk_tipo_producto__codigo=CODIGO_INSUMO
            )
            nombre = producto.nombre

            producto.is_active = False
            producto.save()
            producto.variantes.update(is_active=False)

            messages.success(request, f'✅ Insumo "{nombre}" eliminado correctamente.')
        except Exception as e:
            messages.error(request, f'❌ Error al eliminar el insumo: {e}')

    return redirect('lista_insumos')

# ====================================================
#  PRODUCTO
# ====================================================
CODIGO_TERMINADO = 'PROD-TERM'


def obtener_indices_variantes(post, prefix):
    """
    Encuentra los índices de variante realmente enviados en el POST,
    buscando '<prefix>_sku_<n>'. Así no importa si el usuario agregó
    o quitó filas en el navegador: no hay huecos ni desalineación.
    """
    patron = re.compile(rf'^{re.escape(prefix)}_sku_(\d+)$')
    indices = [int(m.group(1)) for k in post.keys() if (m := patron.match(k))]
    return sorted(indices)


def resolver_foto_archivo_o_url(post, files, campo_archivo, campo_url):
    """
    Devuelve (archivo, url). Nunca ambos a la vez:
    si el usuario mandó archivo, ese manda y la url se descarta.
    Si solo mandó URL, no se descarga nada, solo se guarda el texto.
    """
    archivo = files.get(campo_archivo)
    url = post.get(campo_url, '').strip() or None
    if archivo:
        url = None
    return archivo, url


def crear_variantes_desde_form(producto, post, files, prefix='variante'):
    """Crea N variantes nuevas a partir de campos indexados en el form."""
    creadas = []
    for i in obtener_indices_variantes(post, prefix):
        sku = post.get(f'{prefix}_sku_{i}', '').strip()
        if not sku:
            continue  # fila vacía, se ignora

        if ProductoVariante.objects.filter(sku=sku).exists():
            raise ValueError(f"El SKU '{sku}' ya está en uso (fila de variante {i + 1}).")

        nombre_variante = post.get(f'{prefix}_nombre_{i}', '').strip() or f'Variante {i + 1}'

        archivo_foto, url_foto = resolver_foto_archivo_o_url(
            post, files, f'{prefix}_foto_{i}', f'{prefix}_foto_url_{i}'
        )

        variante = ProductoVariante.objects.create(
            producto=producto,
            nombre_variante=nombre_variante,
            sku=sku,
            codigo_barras=post.get(f'{prefix}_codigo_barras_{i}') or None,
            precio_referencial=safe_decimal(post.get(f'{prefix}_precio_{i}')),
            costo=safe_decimal(post.get(f'{prefix}_costo_{i}')),
            maneja_stock=post.get(f'{prefix}_maneja_stock_{i}') == 'on',
            foto=archivo_foto,
            foto_url=url_foto,
        )
        creadas.append(variante)
    return creadas

@login_required
@permiso_requerido('producto_list', 'ver')
def producto_list(request):
    empresa = request.user.sucursal.fk_empresa
    busqueda = request.GET.get('q', '').strip()

    variantes_queryset = (
        ProductoVariante.objects
        .filter(
            is_active=True,
            producto__is_active=True,
            producto__fk_tipo_producto__codigo=CODIGO_TERMINADO,
            producto__fk_empresa=empresa,
        )
        .select_related('producto', 'producto__unidad_medida', 'producto__category')
        .order_by('producto__nombre', 'nombre_variante', 'id')
    )

    if busqueda:
        variantes_queryset = variantes_queryset.filter(
            Q(producto__nombre__icontains=busqueda) |
            Q(nombre_variante__icontains=busqueda) |
            Q(producto__category__name__icontains=busqueda) |
            Q(sku__icontains=busqueda) |
            Q(codigo_barras__icontains=busqueda)
        )

    paginator = Paginator(variantes_queryset, 100)
    page_number = request.GET.get('page')
    variantes = paginator.get_page(page_number)

    context = {
        'variantes': variantes,
        'unidades_medida': UnidadMedida.objects.filter(is_active=True, fk_empresa=empresa),
        'categorias': Category.objects.filter(is_active=True, fk_empresa=empresa),
        'titulo': 'Productos Terminados',
        'paginator': paginator,
        'page_obj': variantes,
        'is_paginated': variantes.has_other_pages(),
        'busqueda': busqueda,
    }
    return render(request, 'inventario/lista_productos.html', context)

@login_required
@permiso_requerido('producto_list', 'crear')
def crear_producto_terminado(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                empresa = request.user.sucursal.fk_empresa
                nombre = request.POST.get('nombre', '').strip()
                unidad_medida_id = request.POST.get('unidad_medida')
                visibilidad = request.POST.get('visibilidad', 'ambos')  # 'venta' | 'compra' | 'ambos'

                if visibilidad not in ('venta', 'compra', 'ambos'):
                    raise ValueError("Selecciona dónde estará visible el producto.")
                if not nombre:
                    raise ValueError("El nombre del producto es obligatorio.")
                if not unidad_medida_id:
                    raise ValueError("La unidad de medida es obligatoria.")

                tipo_terminado = get_tipo_producto(CODIGO_TERMINADO)

                producto = Producto.objects.create(
                    nombre=nombre,
                    descripcion=request.POST.get('descripcion', ''),
                    fk_empresa=empresa,
                    fk_tipo_producto=tipo_terminado,
                    unidad_medida_id=unidad_medida_id,
                    category_id=request.POST.get('categoria') or None,
                    unidades_por_caja=int(request.POST.get('unidades_por_caja', 0) or 0),
                    tara_por_caja=safe_decimal(request.POST.get('tara_por_caja')),
                    visible_venta=visibilidad in ('venta', 'ambos'),
                    visible_compra=visibilidad in ('compra', 'ambos'),
                )

                tiene_variantes = request.POST.get('tiene_variantes') == 'on'

                if tiene_variantes:
                    variantes_creadas = crear_variantes_desde_form(
                        producto, request.POST, request.FILES, prefix='variante'
                    )
                    if not variantes_creadas:
                        raise ValueError("Debes completar al menos una variante con SKU.")
                else:
                    sku = request.POST.get('sku_base', '').strip()
                    if not sku:
                        raise ValueError("El SKU es obligatorio.")
                    if ProductoVariante.objects.filter(sku=sku).exists():
                        raise ValueError(f"El SKU '{sku}' ya está en uso.")

                    archivo_foto, url_foto = resolver_foto_archivo_o_url(
                        request.POST, request.FILES, 'foto_base', 'foto_url_base'
                    )

                    ProductoVariante.objects.create(
                        producto=producto,
                        nombre_variante="Único",
                        sku=sku,
                        codigo_barras=request.POST.get('codigo_barras_base') or None,
                        precio_referencial=safe_decimal(request.POST.get('precio_base')),
                        costo=safe_decimal(request.POST.get('costo_base')),
                        maneja_stock=request.POST.get('maneja_stock_base') == 'on',
                        foto=archivo_foto,
                        foto_url=url_foto,
                    )

                messages.success(request, f'✅ Producto "{nombre}" creado correctamente.')
                return redirect('producto_list')

        except ValueError as e:
            messages.error(request, f'❌ {e}')
        except IntegrityError:
            messages.error(request, '❌ Ya existe un registro con ese SKU.')
        except Exception as e:
            messages.error(request, f'❌ Error al crear el producto: {e}')

    return render(request, 'inventario/crear_producto.html', {
        'unidades_medida': UnidadMedida.objects.filter(is_active=True, fk_empresa=request.user.sucursal.fk_empresa),
        'categorias': Category.objects.filter(is_active=True, fk_empresa=request.user.sucursal.fk_empresa),
    })


@login_required
@permiso_requerido('producto_list', 'editar')
def producto_variante_edit(request, variante_id):
    empresa = request.user.sucursal.fk_empresa
    variante = get_object_or_404(
        ProductoVariante,
        id=variante_id,
        producto__fk_empresa=empresa,
        producto__fk_tipo_producto__codigo=CODIGO_TERMINADO,
    )
    producto = variante.producto

    if request.method == 'POST':
        try:
            with transaction.atomic():
                # ---- Datos del PRODUCTO PADRE (afectan a TODAS sus variantes) ----
                nombre = request.POST.get('nombre', '').strip()
                unidad_medida_id = request.POST.get('unidad_medida')
                visibilidad = request.POST.get('visibilidad', 'ambos')

                if not nombre:
                    raise ValueError("El nombre del producto es obligatorio.")
                if not unidad_medida_id:
                    raise ValueError("La unidad de medida es obligatoria.")
                if visibilidad not in ('venta', 'compra', 'ambos'):
                    raise ValueError("Selecciona dónde estará visible el producto.")

                producto.nombre = nombre
                producto.descripcion = request.POST.get('descripcion', '')
                producto.unidad_medida_id = unidad_medida_id
                producto.category_id = request.POST.get('categoria') or None
                producto.unidades_por_caja = int(request.POST.get('unidades_por_caja', 0) or 0)
                producto.tara_por_caja = safe_decimal(request.POST.get('tara_por_caja'))
                producto.visible_venta = visibilidad in ('venta', 'ambos')
                producto.visible_compra = visibilidad in ('compra', 'ambos')
                producto.save()

                # ---- Datos PROPIOS de esta variante ----
                nuevo_sku = request.POST.get('sku', '').strip()
                if not nuevo_sku:
                    raise ValueError("El SKU es obligatorio.")
                if ProductoVariante.objects.filter(sku=nuevo_sku).exclude(id=variante.id).exists():
                    raise ValueError(f"El SKU '{nuevo_sku}' ya está en uso.")

                variante.nombre_variante = request.POST.get('nombre_variante', '').strip() or variante.nombre_variante
                variante.sku = nuevo_sku
                variante.codigo_barras = request.POST.get('codigo_barras') or None
                variante.precio_referencial = safe_decimal(request.POST.get('precio'))
                variante.costo = safe_decimal(request.POST.get('costo'))
                variante.maneja_stock = request.POST.get('maneja_stock') == 'on'

                archivo_foto, url_foto = resolver_foto_archivo_o_url(
                    request.POST, request.FILES, 'foto', 'foto_url'
                )
                if archivo_foto:
                    variante.foto = archivo_foto
                    variante.foto_url = None
                elif url_foto:
                    variante.foto = None
                    variante.foto_url = url_foto
                # si no mandó ninguno, se respeta la foto actual

                variante.save()

                messages.success(request, f'✅ "{producto.nombre} — {variante.nombre_variante}" actualizado correctamente.')
                return redirect('producto_list')

        except ValueError as e:
            messages.error(request, f'❌ {e}')
        except IntegrityError:
            messages.error(request, '❌ Ya existe un registro con ese SKU.')
        except Exception as e:
            messages.error(request, f'❌ Error al actualizar: {e}')

    context = {
        'variante': variante,
        'producto': producto,
        'unidades_medida': UnidadMedida.objects.filter(is_active=True, fk_empresa=empresa),
        'categorias': Category.objects.filter(is_active=True, fk_empresa=empresa),
        'hermanas_activas': producto.variantes.filter(is_active=True).exclude(id=variante.id).count(),
        'visibilidad_actual': (
            'ambos' if producto.visible_venta and producto.visible_compra
            else ('venta' if producto.visible_venta else 'compra')
        ),
    }
    return render(request, 'inventario/editar_variante.html', context)

@login_required
@permiso_requerido('producto_list', 'eliminar')
def producto_terminado_delete(request):
    if request.method == 'POST':
        try:
            empresa = request.user.sucursal.fk_empresa
            variante_id = request.POST.get('id')
            variante = get_object_or_404(
                ProductoVariante,
                id=variante_id,
                producto__fk_empresa=empresa,
                producto__fk_tipo_producto__codigo=CODIGO_TERMINADO,
            )
            producto = variante.producto
            nombre_mostrar = f'{producto.nombre} — {variante.nombre_variante}'

            variante.is_active = False
            variante.save()

            # Si el producto se quedó sin ninguna variante activa, se apaga también
            if not producto.variantes.filter(is_active=True).exists():
                producto.is_active = False
                producto.save()

            messages.success(request, f'✅ "{nombre_mostrar}" eliminado correctamente.')
        except Exception as e:
            messages.error(request, f'❌ Error al eliminar: {e}')

    return redirect('producto_list')
# ============================================
# API: OBTENER PRODUCTOS GENÉRICOS (EXACTO COMO PHP)
# ============================================
@login_required
def obtener_productos_genericos(request):
    """
    EXACTO como en PHP:
    SELECT DISTINCT p.id, p.nombre
    FROM producto p
    WHERE p.activo = 1
    AND p.tiene_variantes = 1
    AND EXISTS (SELECT 1 FROM producto_variante pv WHERE pv.fk_producto = p.id AND pv.activo = 1)
    """
    productos = Producto.objects.filter(
        is_active=True,
        # En PHP: p.tiene_variantes = 1
        # En Django: filtramos por productos que tienen variantes
    ).filter(
        # EXISTS: que tengan al menos una variante activa
        variantes__is_active=True
    ).distinct().order_by('nombre')
    
    resultados = [{'id': p.id, 'nombre': p.nombre} for p in productos]
    return JsonResponse({'productos': resultados})


# ============================================
# API: OBTENER PRODUCTOS DEFINIDOS (EXACTO COMO PHP)
# ============================================
@login_required
def obtener_productos_definidos(request):
    """
    EXACTO como en PHP:
    SELECT pv.id AS producto_variante_id, p.id AS producto_id, CONCAT(...) AS nombre_completo
    FROM producto_variante pv
    INNER JOIN producto p ON p.id = pv.fk_producto
    WHERE p.activo = 1 AND pv.activo = 1
    """
    variantes = ProductoVariante.objects.filter(
        is_active=True,
        producto__is_active=True
    ).select_related('producto')
    
    resultados = []
    for v in variantes:
        # Excluir packs (como en PHP: p.fk_tipo_producto = 1)
        if v.producto.fk_tipo_producto and v.producto.fk_tipo_producto.codigo == 'COMBO-PACK':
            continue
        
        # Construir nombre_completo como en PHP
        nombre_completo = v.producto.nombre
        if hasattr(v, 'fk_sabor') and v.fk_sabor:
            nombre_completo += f" - {v.fk_sabor.nombre}"
        elif hasattr(v, 'fk_color') and v.fk_color:
            nombre_completo += f" - {v.fk_color.nombre}"
        
        resultados.append({
            'producto_variante_id': v.id,
            'producto_id': v.producto.id,
            'nombre_completo': nombre_completo,
            'sku': v.sku
        })
    
    return JsonResponse({'variantes': resultados})


# ============================================
# LISTAR COMBOS
# ============================================
@login_required
def lista_combos(request):
    combos = Producto.objects.filter(
        is_active=True,
        fk_tipo_producto__codigo='COMBO-PACK'
    ).select_related('category', 'unidad_medida').prefetch_related('variantes__padre_packs')
    
    context = {
        'combos': combos,
        'categorias': Category.objects.filter(is_active=True),
        'unidades_medida': UnidadMedida.objects.filter(is_active=True),
    }
    return render(request, 'inventario/lista_combos.html', context)


# ============================================
# CREAR PACK (EXACTO COMO PHP)
# ============================================
@login_required
def crear_combo(request):
    if request.method != 'POST':
        return redirect('lista_combos')
    
    try:
        with transaction.atomic():
            # ========= 1. VALIDACIONES =========
            nombre = request.POST.get('nombre', '').strip()
            if len(nombre) < 2:
                messages.error(request, 'El nombre debe tener al menos 2 caracteres')
                return redirect('lista_combos')
            
            precio_venta = Decimal(request.POST.get('precio_venta', '0'))
            if precio_venta <= 0:
                messages.error(request, 'El precio de venta debe ser un número positivo')
                return redirect('lista_combos')
            
            precio_costo = Decimal(request.POST.get('precio_costo', '0'))
            if precio_costo < 0:
                messages.error(request, 'El precio de costo no puede ser negativo')
                return redirect('lista_combos')
            
            # ========= 2. RECOLECTAR DETALLES =========
            detalles = []
            i = 0
            while True:
                tipo_key = f'detalle_{i}_tipo'
                if tipo_key not in request.POST:
                    break
                
                tipo = request.POST.get(tipo_key)
                cantidad = Decimal(request.POST.get(f'detalle_{i}_cantidad', '1'))
                precio_unitario = Decimal(request.POST.get(f'detalle_{i}_precio_unitario', '0'))
                
                if cantidad <= 0:
                    raise Exception(f"Cantidad inválida en el detalle #{i+1}")
                
                if tipo == 'variante':
                    # Producto DEFINIDO
                    variante_id = request.POST.get(f'detalle_{i}_producto_variante_id')
                    if variante_id:
                        try:
                            variante = ProductoVariante.objects.get(id=variante_id, is_active=True)
                            detalles.append({
                                'producto_variante_id': variante_id,
                                'producto_item_id': None,
                                'cantidad': cantidad,
                                'precio_unitario': precio_unitario if precio_unitario > 0 else variante.costo
                            })
                        except ProductoVariante.DoesNotExist:
                            raise Exception(f"La variante no existe en el detalle #{i+1}")
                
                elif tipo == 'item':
                    # Producto GENÉRICO
                    producto_id = request.POST.get(f'detalle_{i}_producto_item_id')
                    if producto_id:
                        try:
                            producto = Producto.objects.get(id=producto_id, is_active=True)
                            detalles.append({
                                'producto_variante_id': None,
                                'producto_item_id': producto_id,
                                'cantidad': cantidad,
                                'precio_unitario': precio_unitario
                            })
                        except Producto.DoesNotExist:
                            raise Exception(f"El producto no existe en el detalle #{i+1}")
                
                i += 1
            
            if not detalles:
                raise Exception("El pack debe tener al menos un producto")
            
            # ========= 3. CREAR PRODUCTO PACK =========
            tipo_combo = TipoProducto.objects.get(codigo='COMBO-PACK')
            
            producto = Producto.objects.create(
                nombre=nombre,
                descripcion=request.POST.get('descripcion', ''),
                fk_empresa=request.user.fk_empresa,
                fk_tipo_producto=tipo_combo,
                unidad_medida_id=request.POST.get('unidad_medida'),
                category_id=request.POST.get('categoria') or None,
                is_active=True
            )
            
            # ========= 4. CREAR VARIANTE DEL PACK =========
            codigo = request.POST.get('codigo', '').strip()
            if not codigo:
                codigo = generar_codigo_pack()
            
            if ProductoVariante.objects.filter(sku=codigo).exists():
                raise Exception(f"El código '{codigo}' ya existe")
            
            variante = ProductoVariante.objects.create(
                producto=producto,
                nombre_variante='Única',
                sku=codigo,
                codigo_barras=request.POST.get('codigo_barras', ''),
                precio_referencial=precio_venta,
                costo=precio_costo,
                maneja_stock=False,
                is_active=True
            )
            
            # ========= 5. CREAR DETALLES =========
            for detalle in detalles:
                DetallePack.objects.create(
                    producto_padre=variante,
                    producto_variante_id=detalle['producto_variante_id'],
                    producto_id=detalle['producto_item_id'],
                    cantidad=detalle['cantidad'],
                    costo_unitario=detalle['precio_unitario']
                )
            
            messages.success(request, f'¡Pack "{nombre}" creado exitosamente con {len(detalles)} productos!')
            
    except Exception as e:
        messages.error(request, str(e))
    
    return redirect('lista_combos')


def generar_codigo_pack():
    prefix = "PACK-"
    ultimo = ProductoVariante.objects.filter(sku__startswith=prefix).order_by('-sku').first()
    numero = int(ultimo.sku.split('-')[1]) + 1 if ultimo else 1
    return f"{prefix}{str(numero).zfill(4)}"

#=====================================================
# IMPORTACION DE PRODUCTOS
#=====================================================

CAMPOS_EXCEL = [
    'nombre', 'sku', 'unidad_medida', 'categoria', 'costo',
    'precio_referencial', 'codigo_barras', 'foto_url',
    'unidades_por_caja', 'tara_por_caja', 'maneja_stock',
]
 
def normalizar(valor):
    return '' if valor is None else str(valor).strip()
 
def _bool_desde_texto(valor):
    return normalizar(valor).lower() in ('1', 'si', 'sí', 'true', 'x')

# ============================================================
# PASO 1: subir Excel -> mostrar vista previa editable
# ============================================================
@login_required
@permiso_requerido('producto_list', 'crear')
def importar_productos_terminados(request):
    if request.method == 'POST':
        archivo = request.FILES.get('archivo_excel')
 
        if not archivo:
            messages.error(request, '❌ Debes seleccionar un archivo Excel.')
            return render(request, 'productos/producto_terminado_import.html', {'titulo': 'Importar Productos Terminados'})
 
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            hoja = wb['ProductosTerminados'] if 'ProductosTerminados' in wb.sheetnames else wb.active
        except InvalidFileException:
            messages.error(request, '❌ El archivo no es un Excel válido (.xlsx).')
            return render(request, 'productos/producto_terminado_import.html', {'titulo': 'Importar Productos Terminados'})
 
        filas_preview = []
        for idx, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            if not fila or not any(fila):
                continue
            data = dict(zip(CAMPOS_EXCEL, fila))
            filas_preview.append({
                'fila_excel': idx,
                'nombre': normalizar(data.get('nombre')),
                'sku': normalizar(data.get('sku')),
                'unidad_medida': normalizar(data.get('unidad_medida')),
                'categoria': normalizar(data.get('categoria')),
                'costo': data.get('costo') if data.get('costo') is not None else 0,
                'precio_referencial': data.get('precio_referencial') if data.get('precio_referencial') is not None else 0,
                'codigo_barras': normalizar(data.get('codigo_barras')),
                'foto_url': normalizar(data.get('foto_url')),
                'unidades_por_caja': data.get('unidades_por_caja') if data.get('unidades_por_caja') is not None else 0,
                'tara_por_caja': data.get('tara_por_caja') if data.get('tara_por_caja') is not None else 0,
                'maneja_stock': _bool_desde_texto(data.get('maneja_stock')),
            })
 
        if not filas_preview:
            messages.warning(request, '⚠️ El Excel no tiene filas con datos para revisar.')
            return render(request, 'inventario/producto_terminado_import.html', {'titulo': 'Importar Productos Terminados'})
 
        context = {
            'filas': filas_preview,
            'titulo': 'Revisar antes de importar',
        }
        return render(request, 'inventario/producto_terminado_import_preview.html', context)
 
    return render(request, 'inventario/producto_terminado_import.html', {'titulo': 'Importar Productos Terminados'})
 
 
# ============================================================
# PASO 2: confirmar -> guardar todo lo que quedó en la revisión
# ============================================================
@login_required
@permiso_requerido('producto_list', 'crear')
def importar_productos_terminados_confirmar(request):
    if request.method != 'POST':
        return redirect('importar_productos_terminados')
 
    empresa = request.user.sucursal.fk_empresa
    tipo_terminado = get_tipo_producto(CODIGO_TERMINADO)
 
    patron = re.compile(r'^fila_nombre_(\d+)$')
    indices = sorted(int(m.group(1)) for k in request.POST.keys() if (m := patron.match(k)))
 
    creados = 0
    actualizados = 0
    errores = []
 
    for i in indices:
        try:
            with transaction.atomic():
                nombre = normalizar(request.POST.get(f'fila_nombre_{i}'))
                sku = normalizar(request.POST.get(f'fila_sku_{i}'))
                nombre_unidad = normalizar(request.POST.get(f'fila_unidad_medida_{i}'))
                nombre_categoria = normalizar(request.POST.get(f'fila_categoria_{i}'))
                foto_url = normalizar(request.POST.get(f'fila_foto_url_{i}'))
 
                if not nombre:
                    raise ValueError("Falta el nombre.")
                if not sku:
                    raise ValueError("Falta el SKU.")
                if not nombre_unidad:
                    raise ValueError("Falta la unidad de medida.")
 
                unidad = UnidadMedida.objects.filter(nombre__iexact=nombre_unidad, fk_empresa=empresa).first()
                if not unidad:
                    unidad = UnidadMedida.objects.create(nombre=nombre_unidad, fk_empresa=empresa)
 
                categoria = None
                if nombre_categoria:
                    categoria = Category.objects.filter(name__iexact=nombre_categoria, fk_empresa=empresa).first()
                    if not categoria:
                        categoria = Category.objects.create(name=nombre_categoria, fk_empresa=empresa)
 
                costo = safe_decimal(request.POST.get(f'fila_costo_{i}'))
                precio = safe_decimal(request.POST.get(f'fila_precio_referencial_{i}'))
                unidades_caja = int(request.POST.get(f'fila_unidades_por_caja_{i}') or 0)
                tara_caja = safe_decimal(request.POST.get(f'fila_tara_por_caja_{i}'))
                codigo_barras = normalizar(request.POST.get(f'fila_codigo_barras_{i}')) or None
                maneja_stock = request.POST.get(f'fila_maneja_stock_{i}') == 'on'
 
                variante_existente = ProductoVariante.objects.filter(sku=sku).first()
 
                if variante_existente:
                    producto = variante_existente.producto
                    if producto.fk_tipo_producto_id != tipo_terminado.id:
                        raise ValueError(f"El SKU '{sku}' ya existe pero pertenece a otro tipo de producto.")
 
                    producto.nombre = nombre
                    producto.unidad_medida = unidad
                    producto.category = categoria
                    producto.unidades_por_caja = unidades_caja
                    producto.tara_por_caja = tara_caja
                    producto.save()
 
                    variante = variante_existente
                else:
                    producto = Producto.objects.create(
                        nombre=nombre,
                        fk_empresa=empresa,
                        fk_tipo_producto=tipo_terminado,
                        unidad_medida=unidad,
                        category=categoria,
                        unidades_por_caja=unidades_caja,
                        tara_por_caja=tara_caja,
                    )
                    variante = ProductoVariante(producto=producto, nombre_variante="Único", sku=sku)
 
                variante.sku = sku
                variante.costo = costo
                variante.precio_referencial = precio
                variante.codigo_barras = codigo_barras
                variante.maneja_stock = maneja_stock
                variante.is_active = True
 
                # Foto por URL -> se descarga y se guarda como archivo real.
                # Si falla, seguimos sin foto (no rompe la fila completa).
                if foto_url:
                    try:
                        import requests
                        resp = requests.get(foto_url, timeout=6)
                        if resp.status_code == 200 and resp.content:
                            nombre_archivo = foto_url.split('/')[-1].split('?')[0] or f'{sku}.jpg'
                            variante.foto.save(nombre_archivo, ContentFile(resp.content), save=False)
                    except Exception:
                        pass
 
                variante.save()
 
                if variante_existente:
                    actualizados += 1
                else:
                    creados += 1
 
        except Exception as e:
            errores.append({
                'fila': i + 2,
                'sku': normalizar(request.POST.get(f'fila_sku_{i}')),
                'motivo': str(e),
            })
 
    if errores:
        messages.warning(
            request,
            f"⚠️ Importación con errores: {creados} creados, {actualizados} actualizados, {len(errores)} con error."
        )
    else:
        messages.success(request, f"✅ Importación completa: {creados} creados, {actualizados} actualizados.")
 
    context = {'errores': errores, 'creados': creados, 'actualizados': actualizados, 'titulo': 'Resultado de la importación'}
    return render(request, 'inventario/producto_terminado_import_resultado.html', context)

# ====================================================
#  PRECIO PRODUCTO
# ====================================================

@login_required
@permiso_requerido('precioproducto_list', 'ver')
def registrar_precios_por_sucursal(request):
    
    sucursal_id = request.GET.get('sucursal_id')
    canal_id = request.GET.get('canal_id')
    
    # Listas para los selects
    sucursales = Sucursal.objects.filter(estado=True, fk_empresa=request.user.sucursal.fk_empresa).order_by('nombre')
    canales = CanalVenta.objects.filter(is_active=True, fk_empresa=request.user.sucursal.fk_empresa).order_by('nombre')

    # IMPORTANTE: Ahora iteramos sobre Variantes, no sobre Productos genéricos
    variantes = []
    precios_dict = {}

    if sucursal_id:
        # Traemos las variantes con el nombre del producto relacionado para mostrarlo en la tabla
        variantes = ProductoVariante.objects.filter(
            producto__is_active=True, 
            producto__fk_empresa=request.user.sucursal.fk_empresa
        ).select_related('producto').order_by('producto__nombre', 'nombre_variante')

        precios_existentes = PrecioProducto.objects.filter(sucursal_id=sucursal_id, activo=True)
        
        if canal_id:
            precios_existentes = precios_existentes.filter(canal_id=canal_id)

        # Diccionario indexado por (variante_id, canal_id)
        precios_dict = {(p.producto_variante_id, p.canal_id): p for p in precios_existentes}

        if request.method == 'POST':
            fecha_nueva = request.POST.get('fecha')
            if fecha_nueva and 'T' in fecha_nueva:
                fecha_nueva = fecha_nueva.split('T')[0]

            try:
                with transaction.atomic():
                    for var in variantes:
                        for canal in canales:
                            if canal_id and str(canal.id) != canal_id:
                                continue

                            # El ID del input en el HTML debe ser precio_VARID_CANALID
                            precio_input = request.POST.get(f'precio_{var.id}_{canal.id}')
                            
                            if precio_input:
                                try:
                                    precio_valor = Decimal(precio_input.replace(',', '.'))
                                except (InvalidOperation, ValueError):
                                    continue # O manejar error de formato

                                key = (var.id, canal.id)
                                
                                # 1. Desactivar el precio anterior de esta VARIANTE en esta SUCURSAL y CANAL
                                PrecioProducto.objects.filter(
                                    producto_variante_id=var.id,
                                    sucursal_id=sucursal_id,
                                    canal_id=canal.id,
                                    activo=True
                                ).update(activo=False)

                                # 2. Crear el nuevo precio histórico
                                PrecioProducto.objects.create(
                                    producto_variante_id=var.id, # CAMBIO CLAVE
                                    sucursal_id=sucursal_id,
                                    canal_id=canal.id,
                                    fecha=fecha_nueva or timezone.now(),
                                    precio=precio_valor,
                                    activo=True
                                )

                messages.success(request, "Precios actualizados correctamente.")
                return redirect(f'{request.path}?sucursal_id={sucursal_id}&canal_id={canal_id}')

            except Exception as e:
                messages.error(request, f"Error al guardar precios: {e}")

    context = {
        'sucursales': sucursales,
        'canales': canales,
        'sucursal_id': sucursal_id,
        'canal_id': canal_id,
        'variantes': variantes, # Pasamos variantes a la tabla
        'precios_dict': precios_dict,
        'fecha_actual': timezone.now(),
    }
    return render(request, 'inventario/precios_sucursales.html', context)

# ====================================================
#  TIPO INGRESO
# ====================================================
@login_required
@permiso_requerido('tiposingreso_list', 'ver')
def tiposingreso_list(request):
    tipos = TipoIngreso.objects.filter(is_active=True, fk_empresa=request.user.sucursal.fk_empresa).order_by('-created_at')
    return render(request, 'inventario/lista_tiposingreso.html', {'tipos': tipos})

@login_required
@permiso_requerido('tiposingreso_list', 'crear')
def tiposingreso_create(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        is_active = True

        # 🔥 Validación de nombre duplicado en activos
        if TipoIngreso.objects.filter(nombre__iexact=nombre, is_active=True, fk_empresa=request.user.sucursal.fk_empresa).exists():
            messages.error(request, f'Ya existe un tipo de ingreso activo con el nombre "{nombre}".')
            return redirect('tiposingreso_list')

        TipoIngreso.objects.create(
            nombre=nombre,
            is_active=is_active,
            fk_empresa=request.user.sucursal.fk_empresa
        )

        messages.success(request, 'Tipo de ingreso creado correctamente.')
        return redirect('tiposingreso_list')

    return redirect('tiposingreso_list')

@login_required
@permiso_requerido('tiposingreso_list', 'editar')
def tiposingreso_edit(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        tipo = get_object_or_404(TipoIngreso, pk=id, fk_empresa=request.user.sucursal.fk_empresa)

        nombre = request.POST.get('nombre', '').strip()
        is_active = True

        # 🔥 Validar nombre duplicado (excluyendo el mismo tipo)
        if TipoIngreso.objects.filter(nombre__iexact=nombre, is_active=True, fk_empresa=request.user.sucursal.fk_empresa).exclude(id=tipo.id).exists():
            messages.error(request, f'Ya existe otro tipo de ingreso activo con el nombre "{nombre}".')
            return redirect('tiposingreso_list')

        tipo.nombre = nombre
        tipo.is_active = is_active
        tipo.save()

        messages.success(request, 'Tipo de ingreso actualizado correctamente.')
        return redirect('tiposingreso_list')

    return redirect('tiposingreso_list')

@login_required
@permiso_requerido('tiposingreso_list', 'eliminar')
def tiposingreso_delete(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        tipo = get_object_or_404(TipoIngreso, pk=id, fk_empresa=request.user.sucursal.fk_empresa)

        tipo.is_active = False
        tipo.save()

        messages.success(request, 'Tipo de ingreso eliminado correctamente.')
        return redirect('tiposingreso_list')

    return redirect('tiposingreso_list')

# ====================================================
#  INGRESO (MAESTRO-DETALLE UNIFICADO)
# ====================================================
@login_required
@permiso_requerido('ingreso_list', 'ver')
def ingreso_list(request):
    ingresos = Ingreso.objects.select_related('tipo', 'usuario', 'sucursal').all().order_by('-created_at')
    return render(request, 'ingreso/list.html', {'ingresos': ingresos})

@login_required
@permiso_requerido('ingreso_list', 'crear')
@transaction.atomic
def ingreso_create(request):
    tipos = TipoIngreso.objects.filter(is_active=True)
    sucursales = Sucursal.objects.filter(estado=True)
    productos = Producto.objects.filter(is_active=True)
    if request.method == 'POST':
        # Maestro
        tipo_id = request.POST.get('tipo')
        usuario_id = request.POST.get('usuario') or request.user.id
        sucursal_id = request.POST.get('sucursal')
        fecha = request.POST.get('fecha')
        descuento_maestro = float(request.POST.get('descuento') or 0)
        observaciones = request.POST.get('observaciones') or ''
        ingreso = Ingreso.objects.create(
            tipo_id=tipo_id,
            usuario_id=usuario_id,
            sucursal_id=sucursal_id,
            fecha=fecha,
            descuento=descuento_maestro,
            observaciones=observaciones,
            total=0  # se actualizará con la sumatoria
        )

        # Detalles (listas)
        productos_list = request.POST.getlist('producto')
        cantidades = request.POST.getlist('cantidad')
        precios = request.POST.getlist('precio')
        descuentos = request.POST.getlist('descuento')  # opcional

        total = 0
        for i, prod_id in enumerate(productos_list):
            if not prod_id:
                continue
            cantidad = float(cantidades[i]) if i < len(cantidades) and cantidades[i] else 0
            precio = float(precios[i]) if i < len(precios) and precios[i] else 0
            descuento_det = float(descuentos[i]) if i < len(descuentos) and descuentos[i] else 0
            subtotal = (cantidad * precio) - descuento_det
            DetalleIngreso.objects.create(
                ingreso=ingreso,
                producto_id=prod_id,
                cantidad=cantidad,
                precio=precio,
                subtotal=subtotal,
                descuento=descuento_det
            )
            total += subtotal

            # Actualizar Stock: sumar cantidad al stock existente (si existe)
            stock, created = Stock.objects.get_or_create(producto_id=prod_id, sucursal_id=sucursal_id, defaults={
                'cantidad_actual': cantidad,
                'cajas_actual': 0,
                'peso_neto_total': 0,
                'costo_unitario_promedio': precio,
                'valor_total': cantidad * precio
            })
            if not created:
                # simple agregado de cantidad y recalculo valor_total y CPP básico (puedes cambiar por CPP exacto)
                prev_total_qty = float(stock.cantidad_actual or 0)
                prev_val_total = float(stock.valor_total or 0)
                new_qty = prev_total_qty + cantidad
                new_val_total = prev_val_total + (cantidad * precio)
                stock.cantidad_actual = new_qty
                stock.valor_total = new_val_total
                # costo_unitario_promedio (simple promedio ponderado)
                stock.costo_unitario_promedio = (new_val_total / new_qty) if new_qty else 0
                stock.save()

        ingreso.total = total - descuento_maestro
        ingreso.save()

        messages.success(request, 'Ingreso y sus detalles registrados correctamente.')
        return redirect('ingreso_list')

    # GET
    usuarios = Usuario.objects.filter(is_active=True)
    return render(request, 'ingreso/create.html', {
        'tipos': tipos,
        'sucursales': sucursales,
        'productos': productos,
        'usuarios': usuarios
    })

@login_required
@permiso_requerido('ingreso_list', 'eliminar')
def ingreso_delete(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        ingreso = get_object_or_404(Ingreso, pk=id)
        ingreso.is_active = False
        ingreso.save()
        # opcional: also deactivate details
        DetalleIngreso.objects.filter(ingreso=ingreso).update(is_active=False)
        messages.success(request, 'Ingreso desactivado correctamente.')
    return redirect('ingreso_list')


# ====================================================
#  TURNO
# ====================================================
@login_required
@permiso_requerido('turno_list', 'ver')
def turno_list(request):
    turnos = Turno.objects.filter(is_active=True, fk_empresa=request.user.sucursal.fk_empresa).order_by('-created_at')
    return render(request, 'inventario/lista_turnos.html', {'turnos': turnos})

@login_required
@permiso_requerido('turno_list', 'crear')
def turno_create(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        hora_inicio = request.POST.get('hora_inicio')
        hora_fin = request.POST.get('hora_fin')

        # Validación simple de campos vacíos
        if not nombre or not hora_inicio or not hora_fin:
            messages.error(request, 'Todos los campos son obligatorios.')
            return redirect('turno_list')

        # Validación nombre duplicado en activos
        if Turno.objects.filter(nombre__iexact=nombre, is_active=True, fk_empresa=request.user.sucursal.fk_empresa).exists():
            messages.error(request, f'Ya existe un turno activo con el nombre "{nombre}".')
            return redirect('turno_list')

        # Validar rango de horas
        if hora_inicio >= hora_fin:
            messages.error(request, 'La hora de inicio debe ser menor a la hora de fin.')
            return redirect('turno_list')

        Turno.objects.create(
            nombre=nombre,
            hora_inicio=hora_inicio,
            hora_fin=hora_fin,
            is_active=True,
            fk_empresa=request.user.sucursal.fk_empresa
        )

        messages.success(request, 'Turno creado correctamente.')
        return redirect('turno_list')

    return redirect('turno_list')

@login_required
@permiso_requerido('turno_list', 'editar')
def turno_edit(request):
    
    if request.method == 'POST':
        id = request.POST.get('id')
        turno = get_object_or_404(Turno, pk=id, fk_empresa=request.user.sucursal.fk_empresa)

        nombre = request.POST.get('nombre', '').strip()
        hora_inicio = request.POST.get('hora_inicio')
        hora_fin = request.POST.get('hora_fin')

        if not nombre or not hora_inicio or not hora_fin:
            messages.error(request, 'Todos los campos son obligatorios.')
            return redirect('turno_list')

        # Validar duplicado EXCLUYENDO el mismo turno
        if Turno.objects.filter(nombre__iexact=nombre, is_active=True, fk_empresa=request.user.sucursal.fk_empresa).exclude(id=turno.id).exists():
            messages.error(request, f'Ya existe otro turno activo con el nombre "{nombre}".')
            return redirect('turno_list')

        # Validar rango
        if hora_inicio >= hora_fin:
            messages.error(request, 'La hora de inicio debe ser menor a la hora de fin.')
            return redirect('turno_list')

        turno.nombre = nombre
        turno.hora_inicio = hora_inicio
        turno.hora_fin = hora_fin
        turno.save()

        messages.success(request, 'Turno actualizado correctamente.')
        return redirect('turno_list')

    return redirect('turno_list')

@login_required
@permiso_requerido('turno_list', 'eliminar')
def turno_delete(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        turno = get_object_or_404(Turno, pk=id, fk_empresa=request.user.sucursal.fk_empresa)

        turno.is_active = False
        turno.save()

        messages.success(request, 'Turno desactivado correctamente.')
        return redirect('turno_list')

    return redirect('turno_list')

# ====================================================
#  CAJA
# ====================================================
@login_required
@permiso_requerido('caja_list', 'ver')
def caja_list(request):
    empresa = request.user.sucursal.fk_empresa
    cajas = Caja.objects.filter(is_active=True, fk_empresa=empresa).order_by('-created_at')
    sucursales = Sucursal.objects.filter(estado=True, fk_empresa=empresa)
    
    return render(request, 'ventas/lista_cajas.html', {
        'cajas': cajas, 
        'sucursales': sucursales
    })

@login_required
@permiso_requerido('caja_list', 'crear')
def caja_create(request):
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        nombre = request.POST.get('nombre', '').strip()
        sucursal_id = request.POST.get('sucursal')
        saldo_inicial = request.POST.get('saldo_inicial') or 0

        # VALIDACIÓN 1: Nombre vacío
        if not nombre:
            messages.error(request, 'El nombre de la caja es obligatorio.')
            return redirect('caja_list')

        # VALIDACIÓN 2: Duplicado en la misma empresa
        if Caja.objects.filter(nombre__iexact=nombre, is_active=True, fk_empresa=empresa).exists():
            messages.error(request, f'Ya existe una caja llamada "{nombre}".')
            return redirect('caja_list')

        # VALIDACIÓN 3: Seguridad de Sucursal (Que pertenezca a la empresa)
        if not Sucursal.objects.filter(id=sucursal_id, fk_empresa=empresa).exists():
            messages.error(request, 'La sucursal seleccionada no es válida.')
            return redirect('caja_list')

        Caja.objects.create(
            nombre=nombre,
            sucursal_id=sucursal_id,
            saldo_inicial=saldo_inicial,
            is_active=True,
            fk_empresa=empresa
        )
        messages.success(request, 'Caja creada correctamente.')
        
    return redirect('caja_list')

@login_required
@permiso_requerido('caja_list', 'editar')
def caja_edit(request):
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        id = request.POST.get('id')
        
        # Seguridad: Solo editamos si pertenece a la empresa
        caja = get_object_or_404(Caja, pk=id, fk_empresa=empresa)
        
        nombre = request.POST.get('nombre', '').strip()
        sucursal_id = request.POST.get('sucursal')
        saldo_inicial = request.POST.get('saldo_inicial')

        # VALIDACIÓN 1: Nombre vacío
        if not nombre:
            messages.error(request, 'El nombre no puede estar vacío.')
            return redirect('caja_list')

        # VALIDACIÓN 2: Duplicado (excluyendo la caja actual)
        if Caja.objects.filter(nombre__iexact=nombre, is_active=True, fk_empresa=empresa).exclude(id=caja.id).exists():
            messages.error(request, f'El nombre "{nombre}" ya está siendo usado por otra caja.')
            return redirect('caja_list')

        # VALIDACIÓN 3: Seguridad de Sucursal
        if not Sucursal.objects.filter(id=sucursal_id, fk_empresa=empresa).exists():
            messages.error(request, 'Sucursal no válida.')
            return redirect('caja_list')

        caja.nombre = nombre
        caja.sucursal_id = sucursal_id
        if saldo_inicial is not None:
            caja.saldo_inicial = saldo_inicial
            
        caja.save()
        messages.success(request, 'Caja actualizada correctamente.')
        
    return redirect('caja_list')

@login_required
@permiso_requerido('caja_list', 'eliminar')
def caja_delete(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        empresa = request.user.sucursal.fk_empresa
        caja = get_object_or_404(Caja, pk=id, fk_empresa=empresa)
        
        caja.is_active = False
        caja.save()
        messages.success(request, 'Caja eliminada correctamente.')
        
    return redirect('caja_list')

# ====================================================
#  CAJA TURNO
# ====================================================
@login_required
def caja_turno_list(request):
    caja_turnos = CajaTurno.objects.select_related('caja', 'turno', 'usuario').all().order_by('-created_at')
    return render(request, 'caja_turno/list.html', {'caja_turnos': caja_turnos})

@login_required
def abrir_caja(request):
    usuario = request.user

    # Verificar si ya tiene una caja abierta
    caja_abierta = CajaTurno.objects.filter(
        usuario=usuario,
        sucursal=usuario.sucursal,
        estado='ABIERTA'
    ).first()

    if caja_abierta:
        messages.warning(
            request,
            'Ya tienes una caja abierta. Debes cerrarla antes de abrir otra.'
        )
        return redirect('crear_venta')

    cajas = Caja.objects.filter(is_active=True, fk_empresa=usuario.sucursal.fk_empresa).order_by('-created_at')
    turnos = Turno.objects.filter(is_active=True,fk_empresa=usuario.sucursal.fk_empresa).order_by('-created_at')

    if request.method == 'POST':

        caja_id = request.POST.get('caja')
        turno_id = request.POST.get('turno')
        cajachica_apertura = request.POST.get('cajachica_apertura') or 0
        observaciones_apertura = request.POST.get('observaciones_apertura', '')

        caja_turno = CajaTurno.objects.create(
            caja_id=caja_id,
            turno_id=turno_id,
            usuario=usuario,
            sucursal=usuario.sucursal,
            cajachica_apertura=cajachica_apertura,
            observaciones_apertura=observaciones_apertura,
            estado='ABIERTA',
            fecha_apertura=timezone.now()
        )

        MovimientoCaja.objects.create(
            caja_turno=caja_turno,
            tipo='APERTURA',
            monto=cajachica_apertura,
            descripcion='Apertura de caja',
            usuario=usuario
        )

        messages.success(request, 'Caja abierta correctamente.')
        return redirect('crear_venta')

    context = {
        'cajas': cajas,
        'turnos': turnos,
        'usuario': usuario,
        'sucursal': usuario.sucursal,
        'fecha_apertura': timezone.now(),
    }

    return render(request, 'ventas/abrir_caja.html', context)

@login_required
def cerrar_caja(request):

    usuario = request.user

    caja_turno = CajaTurno.objects.filter(
        usuario=usuario,
        sucursal=usuario.sucursal,
        estado='ABIERTA',
        is_active=True
    ).first()

    if not caja_turno:
        messages.warning(request, 'No tienes una caja abierta.')
        return redirect('abrir_caja')

    if request.method == 'POST':

        caja_turno.monto_efectivo = Decimal(
            request.POST.get('monto_efectivo') or '0'
        )

        caja_turno.monto_qr = Decimal(
            request.POST.get('monto_qr') or '0'
        )

        caja_turno.monto_tarjeta = Decimal(
            request.POST.get('monto_tarjeta') or '0'
        )

        caja_turno.monto_cierre = Decimal(
            request.POST.get('monto_cierre') or '0'
        )

        caja_turno.saldo_teorico = Decimal(
            request.POST.get('saldo_teorico') or '0'
        )

        caja_turno.diferencia = Decimal(
            request.POST.get('diferencia') or '0'
        )

        caja_turno.observaciones_cierre = request.POST.get(
            'observaciones_cierre', ''
        )

        caja_turno.estado = 'CERRADA'
        caja_turno.fecha_cierre = timezone.now()

        caja_turno.save()

        MovimientoCaja.objects.create(
            caja_turno=caja_turno,
            tipo='CIERRE',
            monto=caja_turno.monto_cierre,
            descripcion='Cierre de caja',
            usuario=usuario
        )

        messages.success(request, 'Caja cerrada correctamente.')
        return redirect('abrir_caja')

    return render(
        request,
        'ventas/cerrar_caja.html',
        {
            'caja_turno': caja_turno
        }
    )

@login_required
def comprobante_cierre_caja(request, caja_turno_id):
    """Muestra el comprobante de cierre de caja"""
    caja_turno = get_object_or_404(
        CajaTurno, 
        id=caja_turno_id,
        usuario=request.user,
        sucursal=request.user.sucursal
    )
    
    # Obtener ventas del turno
    ventas = caja_turno.ventas.filter(is_active=True)
    total_ventas = ventas.count()
    total_monto = ventas.aggregate(total=models.Sum('total'))['total'] or 0
    
    # Calcular efectivo esperado (suma de ventas en efectivo)
    total_efectivo_esperado = 0
    for venta in ventas:
        for pago in venta.pagos.all():
            if pago.metodo_pago.nombre.lower() == 'efectivo':
                total_efectivo_esperado += float(pago.monto)
    
    context = {
        'caja_turno': caja_turno,
        'ventas': ventas,
        'total_ventas': total_ventas,
        'total_monto': total_monto,
        'total_efectivo_esperado': total_efectivo_esperado,
        'fecha_cierre': timezone.now(),
    }
    
    return render(request, 'ventas/comprobante_cierre_caja.html', context)

@login_required
def total_efectivo_esperado(request):
    """Retorna el total de efectivo esperado del turno actual"""
    caja_turno = CajaTurno.objects.filter(
        usuario=request.user,
        sucursal=request.user.sucursal,
        is_active=True,
        fecha_cierre__isnull=True
    ).first()
    
    if not caja_turno:
        return JsonResponse({'ok': True, 'total': 0})
    
    from django.db.models import Sum
    total_efectivo = 0
    for venta in caja_turno.ventas.filter(is_active=True):
        for pago in venta.pagos.all():
            if pago.metodo_pago.nombre.lower() == 'efectivo':
                total_efectivo += float(pago.monto)
    
    return JsonResponse({'ok': True, 'total': total_efectivo})

# ====================================================
# REPORTE DE CAJAS
# ====================================================
@login_required
@permiso_requerido('reporte_cajas', 'ver')
def reporte_cajas(request):
    """Vista para el reporte de movimientos de caja"""
    
    usuario = request.user
    empresa = usuario.fk_empresa
    sucursal = usuario.sucursal
    
    # Obtener filtros
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    caja_id = request.GET.get('caja_id')
    turno_id = request.GET.get('turno_id')
    usuario_id = request.GET.get('usuario_id')
    tipo_movimiento = request.GET.get('tipo_movimiento')
    busqueda = request.GET.get('buscar', '')
    
    # Fechas por defecto (hoy)
    hoy = timezone.now().date()
    if not fecha_desde:
        fecha_desde = hoy
    if not fecha_hasta:
        fecha_hasta = hoy
    
    # Convertir fechas a datetime
    try:
        if isinstance(fecha_desde, str):
            fecha_desde_dt = datetime.strptime(str(fecha_desde), '%Y-%m-%d').date()
        else:
            fecha_desde_dt = fecha_desde
    except:
        fecha_desde_dt = hoy
    
    try:
        if isinstance(fecha_hasta, str):
            fecha_hasta_dt = datetime.strptime(str(fecha_hasta), '%Y-%m-%d').date()
        else:
            fecha_hasta_dt = fecha_hasta
    except:
        fecha_hasta_dt = hoy
    
    # Base de movimientos (solo de la empresa del usuario)
    movimientos = MovimientoCaja.objects.filter(
        caja_turno__is_active=True,
        caja_turno__caja__is_active=True,
        caja_turno__caja__fk_empresa=empresa,
        caja_turno__sucursal=sucursal
    ).select_related(
        'caja_turno', 
        'caja_turno__caja', 
        'caja_turno__turno', 
        'caja_turno__usuario', 
        'usuario'
    )
    
    # Aplicar filtros de fecha
    if fecha_desde_dt and fecha_hasta_dt:
        fecha_inicio = datetime.combine(fecha_desde_dt, datetime.min.time())
        fecha_fin = datetime.combine(fecha_hasta_dt, datetime.max.time())
        movimientos = movimientos.filter(
            created_at__gte=fecha_inicio,
            created_at__lte=fecha_fin
        )
    
    # Filtros adicionales
    if caja_id:
        movimientos = movimientos.filter(caja_turno__caja_id=caja_id)
    
    if turno_id:
        movimientos = movimientos.filter(caja_turno__turno_id=turno_id)
    
    if usuario_id:
        movimientos = movimientos.filter(usuario_id=usuario_id)
    
    if tipo_movimiento:
        movimientos = movimientos.filter(tipo=tipo_movimiento)
    
    if busqueda:
        movimientos = movimientos.filter(
            Q(referencia__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )
    
    # Ordenar por fecha descendente
    movimientos = movimientos.order_by('-created_at')
    
    # Calcular total de movimientos - USANDO DjangoSum
    total_movimientos = movimientos.aggregate(
        total=DjangoSum('monto')  # ← FIX: usar DjangoSum
    )['total'] or 0
    
    # KPIs
    # Cajas abiertas en la sucursal
    cajas_abiertas = CajaTurno.objects.filter(
        estado='ABIERTA',
        is_active=True,
        sucursal=sucursal,
        caja__fk_empresa=empresa
    ).select_related('caja', 'usuario', 'turno')
    
    # Saldo total en cajas abiertas
    total_saldo = cajas_abiertas.aggregate(
        total=DjangoSum('saldo_teorico')  # ← FIX: usar DjangoSum
    )['total'] or 0
    
    # Ventas del día
    ventas_hoy = Venta.objects.filter(
        fecha__date=hoy,
        is_active=True,
        caja_turno__is_active=True,
        sucursal=sucursal
    )
    total_ventas_dia = ventas_hoy.aggregate(
        total=DjangoSum('total')  # ← FIX: usar DjangoSum
    )['total'] or 0
    total_ventas_count = ventas_hoy.count()
    
    # Egresos monetarios del día
    egresos_hoy = EgresoMonetario.objects.filter(
        fecha__date=hoy,
        is_active=True,
        caja_turno__is_active=True,
        caja_turno__sucursal=sucursal
    )
    total_egresos_dia = egresos_hoy.aggregate(
        total=DjangoSum('monto')  # ← FIX: usar DjangoSum
    )['total'] or 0
    total_egresos_count = egresos_hoy.count()
    
    # Diferencia
    diferencia = total_ventas_dia - total_egresos_dia
    
    # Obtener símbolo de moneda
    simbolo_moneda = empresa.simbolo_moneda if empresa.simbolo_moneda else 'Bs.'
    
    # Contexto para filtros
    context = {
        'movimientos': movimientos,
        'total_movimientos': total_movimientos,
        'total_saldo': total_saldo,
        'total_ventas_dia': total_ventas_dia,
        'total_ventas_count': total_ventas_count,
        'total_egresos_dia': total_egresos_dia,
        'total_egresos_count': total_egresos_count,
        'diferencia': diferencia,
        'cajas_abiertas': cajas_abiertas,
        'cajas': Caja.objects.filter(
            is_active=True,
            sucursal=sucursal,
            fk_empresa=empresa
        ),
        'turnos': Turno.objects.filter(
            is_active=True,
            fk_empresa=empresa
        ),
        'usuarios': Usuario.objects.filter(
            is_active=True,
            sucursal=sucursal
        ),
        'fecha_desde': fecha_desde_dt,
        'fecha_hasta': fecha_hasta_dt,
        'caja_id': caja_id,
        'turno_id': turno_id,
        'usuario_id': usuario_id,
        'tipo_movimiento': tipo_movimiento,
        'busqueda': busqueda,
        'fecha_actual': timezone.now(),
        'mostrando_hoy': not request.GET.get('fecha_desde') and not request.GET.get('fecha_hasta'),
        'simbolo_moneda': simbolo_moneda,
    }
    
    return render(request, 'inventario/reporte_cajas.html', context)# ====================================================
#  TIPO EGRESO
# ====================================================
@login_required
@permiso_requerido('tipo_egreso_list', 'ver')
def tipo_egreso_list(request):
    tipos = TipoEgreso.objects.filter(is_active=True, fk_empresa=request.user.sucursal.fk_empresa).order_by('-created_at')
    return render(request, 'inventario/lista_tiposegreso.html', {'tipos': tipos})

@login_required
@permiso_requerido('tipo_egreso_list', 'crear')
def tipo_egreso_create(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        is_active = True  # siempre activo al crear

        # 🔥 Validación: nombre duplicado entre los activos
        if TipoEgreso.objects.filter(nombre__iexact=nombre, is_active=True, fk_empresa=request.user.sucursal.fk_empresa).exists():
            messages.error(request, f'Ya existe un tipo de egreso activo con el nombre "{nombre}".')
            return redirect('tipo_egreso_list')

        TipoEgreso.objects.create(
            nombre=nombre,
            is_active=is_active,
            fk_empresa=request.user.sucursal.fk_empresa
        )

        messages.success(request, 'Tipo de egreso creado correctamente.')

    return redirect('tipo_egreso_list')

@login_required
@permiso_requerido('tipo_egreso_list', 'editar')
def tipo_egreso_edit(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        tipo = get_object_or_404(TipoEgreso, pk=id, fk_empresa=request.user.sucursal.fk_empresa)

        nombre = request.POST.get('nombre', '').strip()
        is_active = True  # siempre activo al editar

        # 🔥 Validación: nombre duplicado excluyendo el mismo registro
        if TipoEgreso.objects.filter(nombre__iexact=nombre, is_active=True, fk_empresa=request.user.sucursal.fk_empresa).exclude(id=tipo.id).exists():
            messages.error(request, f'Ya existe otro tipo de egreso activo con el nombre "{nombre}".')
            return redirect('tipo_egreso_list')

        tipo.nombre = nombre
        tipo.is_active = is_active
        tipo.save()

        messages.success(request, 'Tipo de egreso actualizado correctamente.')

    return redirect('tipo_egreso_list')

@login_required
@permiso_requerido('tipo_egreso_list', 'eliminar')
def tipo_egreso_delete(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        tipo = get_object_or_404(TipoEgreso, pk=id, fk_empresa=request.user.sucursal.fk_empresa)

        tipo.is_active = False
        tipo.save()

        messages.success(request, 'Tipo de egreso desactivado correctamente.')

    return redirect('tipo_egreso_list')

# ====================================================
#  EGRESO (MAESTRO-DETALLE)
# ====================================================
@login_required
@permiso_requerido('egreso_list', 'ver')
def egreso_list(request):
    egresos = Egreso.objects.select_related('tipo', 'usuario', 'sucursal').all().order_by('-created_at')
    return render(request, 'egreso/list.html', {'egresos': egresos})

@login_required
@permiso_requerido('egreso_list', 'crear')
@transaction.atomic
def egreso_create(request):
    tipos = TipoEgreso.objects.filter(is_active=True)
    sucursales = Sucursal.objects.filter(estado=True)
    productos = Producto.objects.filter(is_active=True)
    caja_turnos = CajaTurno.objects.filter(is_active=True)
    usuarios = Usuario.objects.filter(is_active=True)
    if request.method == 'POST':
        tipo_id = request.POST.get('tipo')
        usuario_id = request.POST.get('usuario') or request.user.id
        sucursal_id = request.POST.get('sucursal')
        caja_turno_id = request.POST.get('caja_turno') or None
        fecha = request.POST.get('fecha')
        descuento_maestro = float(request.POST.get('descuento') or 0)
        observaciones = request.POST.get('observaciones') or ''

        egreso = Egreso.objects.create(
            tipo_id=tipo_id,
            usuario_id=usuario_id,
            sucursal_id=sucursal_id,
            fecha=fecha,
            descuento=descuento_maestro,
            observaciones=observaciones,
            caja_turno_id=caja_turno_id,
            total=0
        )

        productos_list = request.POST.getlist('producto')
        cantidades = request.POST.getlist('cantidad')
        precios = request.POST.getlist('precio')
        descuentos = request.POST.getlist('descuento')

        total = 0
        for i, prod_id in enumerate(productos_list):
            if not prod_id:
                continue
            cantidad = float(cantidades[i]) if i < len(cantidades) and cantidades[i] else 0
            precio = float(precios[i]) if i < len(precios) and precios[i] else 0
            descuento_det = float(descuentos[i]) if i < len(descuentos) and descuentos[i] else 0
            subtotal = (cantidad * precio) - descuento_det
            DetalleEgreso.objects.create(
                egreso=egreso,
                producto_id=prod_id,
                cantidad=cantidad,
                precio=precio,
                subtotal=subtotal,
                descuento=descuento_det
            )
            total += subtotal

            # actualizar stock: restar cantidad
            try:
                stock = Stock.objects.get(producto_id=prod_id, sucursal_id=sucursal_id)
                stock.cantidad_actual = float(stock.cantidad_actual or 0) - cantidad
                stock.valor_total = float(stock.cantidad_actual or 0) * float(stock.costo_unitario_promedio or 0)
                stock.save()
            except Stock.DoesNotExist:
                # Si no existe, crear negativo o saltar
                pass

        egreso.total = total - descuento_maestro
        egreso.save()
        messages.success(request, 'Egreso registrado correctamente.')
        return redirect('egreso_list')

    return render(request, 'egreso/create.html', {
        'tipos': tipos,
        'sucursales': sucursales,
        'productos': productos,
        'caja_turnos': caja_turnos,
        'usuarios': usuarios
    })

@login_required
@permiso_requerido('egreso_list', 'editar')
@transaction.atomic
def egreso_edit(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        egreso = get_object_or_404(Egreso, pk=id)
        egreso.tipo_id = request.POST.get('tipo')
        egreso.sucursal_id = request.POST.get('sucursal')
        egreso.fecha = request.POST.get('fecha')
        egreso.descuento = float(request.POST.get('descuento') or 0)
        egreso.observaciones = request.POST.get('observaciones') or ''
        egreso.save()

        # eliminar y recrear detalles
        DetalleEgreso.objects.filter(egreso=egreso).delete()
        productos_list = request.POST.getlist('producto')
        cantidades = request.POST.getlist('cantidad')
        precios = request.POST.getlist('precio')
        descuentos = request.POST.getlist('descuento')

        total = 0
        for i, prod_id in enumerate(productos_list):
            if not prod_id:
                continue
            cantidad = float(cantidades[i]) if i < len(cantidades) and cantidades[i] else 0
            precio = float(precios[i]) if i < len(precios) and precios[i] else 0
            descuento_det = float(descuentos[i]) if i < len(descuentos) and descuentos[i] else 0
            subtotal = (cantidad * precio) - descuento_det
            DetalleEgreso.objects.create(
                egreso=egreso,
                producto_id=prod_id,
                cantidad=cantidad,
                precio=precio,
                subtotal=subtotal,
                descuento=descuento_det
            )
            total += subtotal
        egreso.total = total - egreso.descuento
        egreso.save()
        messages.success(request, 'Egreso actualizado correctamente.')
        return redirect('egreso_list')
    else:
        id = request.GET.get('id')
        egreso = get_object_or_404(Egreso, pk=id)
        tipos = TipoEgreso.objects.filter(is_active=True)
        sucursales = Sucursal.objects.filter(estado=True)
        detalles = DetalleEgreso.objects.filter(egreso=egreso)
        caja_turnos = CajaTurno.objects.filter(is_active=True)
        usuarios = Usuario.objects.filter(is_active=True)
        return render(request, 'egreso/edit.html', {
            'egreso': egreso, 'tipos': tipos, 'sucursales': sucursales, 'detalles': detalles,
            'caja_turnos': caja_turnos, 'usuarios': usuarios
        })

@login_required
@permiso_requerido('egreso_list', 'eliminar')
def egreso_delete(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        egreso = get_object_or_404(Egreso, pk=id)
        egreso.is_active = False
        egreso.save()
        DetalleEgreso.objects.filter(egreso=egreso).update(is_active=False)
        messages.success(request, 'Egreso desactivado correctamente.')
    return redirect('egreso_list')


# ====================================================
#  PROVEEDOR
# ====================================================
@login_required
@permiso_requerido('proveedor_list', 'ver')
def proveedor_list(request):
    proveedores = (
        Proveedor.objects.select_related('empresa').filter(is_active=True, empresa_id=request.user.sucursal.fk_empresa.id).order_by('-created_at'))
    empresas = Empresa.objects.filter(estado=True)
    return render(request, 'empresa/lista_proveedores.html', {'proveedores': proveedores, 'empresas': empresas})

@login_required
@permiso_requerido('proveedor_list', 'crear')
def proveedor_create(request):
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        nombre = request.POST.get('nombre', '').strip()
        email = request.POST.get('email', '').strip()
        
        if not nombre:
            messages.error(request, 'El nombre del proveedor es obligatorio.')
            return redirect('proveedor_list')
        if Proveedor.objects.filter(nombre__iexact=nombre, empresa=empresa, is_active=True).exists():
            messages.error(request, f'Ya existe un proveedor llamado "{nombre}" en tu empresa.')
            return redirect('proveedor_list')

        Proveedor.objects.create(
            nombre=nombre,
            contacto=request.POST.get('contacto'),
            telefono=request.POST.get('telefono'),
            email=email,
            direccion=request.POST.get('direccion'),
            empresa=empresa # Asignación automática segura
        )

        messages.success(request, 'Proveedor creado correctamente.')
    return redirect('proveedor_list')

@login_required
@permiso_requerido('proveedor_list', 'editar')
def proveedor_edit(request):
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        id = request.POST.get('id')
        p = get_object_or_404(Proveedor, pk=id, empresa=empresa)

        nombre = request.POST.get('nombre', '').strip()
        
        if Proveedor.objects.filter(nombre__iexact=nombre, empresa=empresa, is_active=True).exclude(id=p.id).exists():
            messages.error(request, f'Ya tienes otro proveedor con el nombre "{nombre}".')
            return redirect('proveedor_list')

        p.nombre = nombre
        p.contacto = request.POST.get('contacto')
        p.telefono = request.POST.get('telefono')
        p.email = request.POST.get('email')
        p.direccion = request.POST.get('direccion')
        p.save()

        messages.success(request, 'Proveedor actualizado correctamente.')
    return redirect('proveedor_list')

@login_required
@permiso_requerido('proveedor_list', 'eliminar')
def proveedor_delete(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        p = get_object_or_404(Proveedor, pk=id)
        p.is_active = False
        p.save()

        messages.success(request, 'Proveedor desactivado correctamente.')

    return redirect('proveedor_list')

# ====================================================
#  COMPRA (MAESTRO-DETALLE)
# ====================================================

def safe_decimal(value):
    if value is None or value == '':
        return Decimal('0.0')
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError):
        return Decimal('0.0')

@login_required
@permiso_requerido('compra_list', 'ver')
def lista_compras(request):
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    sucursal_id = request.GET.get('sucursal_id')
    proveedor_id = request.GET.get('proveedor_id')
    usuario_id = request.GET.get('usuario_id')
    estado_compra = request.GET.get('estado_compra')

    compras = Compra.objects.filter(is_active=True, sucursal=request.user.sucursal).order_by('-fecha').prefetch_related('detallecompra_set')

    # FILTROS
    if fecha_desde:
        try:
            compras = compras.filter(fecha__date__gte=timezone.datetime.strptime(fecha_desde, '%Y-%m-%d').date())
        except:
            pass
    else:
        compras = compras.filter(fecha__date=timezone.now().date())

    if fecha_hasta:
        try:
            compras = compras.filter(fecha__date__lte=timezone.datetime.strptime(fecha_hasta, '%Y-%m-%d').date())
        except:
            pass

    if sucursal_id and sucursal_id.isdigit():
        compras = compras.filter(sucursal_id=int(sucursal_id))

    if proveedor_id and proveedor_id.isdigit():
        compras = compras.filter(proveedor_id=int(proveedor_id))

    if usuario_id and usuario_id.isdigit():
        compras = compras.filter(usuario_id=int(usuario_id))

    if estado_compra in ['0', '1']:
        compras = compras.filter(is_active=bool(int(estado_compra)))

    sucursales = Sucursal.objects.filter(estado=True, fk_empresa=request.user.sucursal.fk_empresa)
    usuarios = Usuario.objects.filter(is_active=True, sucursal=request.user.sucursal.fk_empresa.id)
    proveedores = Proveedor.objects.filter(is_active=True, empresa=request.user.sucursal.fk_empresa)

    return render(request, 'inventario/lista_compras.html', {
        'compras': compras,
        'sucursales': sucursales,
        'usuarios': usuarios,
        'proveedores': proveedores,
        'fecha_desde': fecha_desde or timezone.now().date().strftime('%Y-%m-%d'),
        'fecha_hasta': fecha_hasta or '',
        'sucursal_id': sucursal_id or '',
        'usuario_id': usuario_id or '',
        'proveedor_id': proveedor_id or '',
        'estado_compra': estado_compra or '',
    })


@login_required
def almacenes_por_sucursal(request):
    sucursal_id = request.GET.get('sucursal')

    if not sucursal_id:
        return JsonResponse({'almacenes': []})

    empresa = request.user.sucursal.fk_empresa

    almacenes = Almacen.objects.filter(
        is_active=True,
        sucursal_id=sucursal_id,
        sucursal__fk_empresa=empresa
    ).values('id', 'nombre')

    return JsonResponse({'almacenes': list(almacenes)})

@login_required
def buscar_variantes(request):
    """
    Búsqueda de variantes para autocompletado (compras, traspasos, etc.).
    No devuelve nada hasta 2+ caracteres, y limita a 15 resultados -
    esto es lo que hace viable buscar entre miles de productos sin
    tirar un <select> gigante al navegador.
    """
    termino = request.GET.get('q', '').strip()
    empresa = request.user.fk_empresa
 
    if len(termino) < 2:
        return JsonResponse({'resultados': []})
 
    variantes = (
        ProductoVariante.objects
        .filter(is_active=True, producto__fk_empresa=empresa)
        .filter(
            Q(producto__nombre__icontains=termino) |
            Q(nombre_variante__icontains=termino) |
            Q(sku__icontains=termino) |
            Q(codigo_barras__icontains=termino)
        )
        .select_related('producto')[:15]
    )
 
    resultados = [{
        'id': v.id,
        'texto': f"{v.producto.nombre} - {v.nombre_variante}",
        'sku': v.sku,
        'costo': str(v.costo),
    } for v in variantes]
 
    return JsonResponse({'resultados': resultados})

@login_required
@permiso_requerido('compra_list', 'crear')
def crear_compra(request):

    if request.method == 'POST':
        try:
            with transaction.atomic():
                usuario = request.user
                sucursal_id = request.POST.get('sucursal')
                proveedor_id = request.POST.get('proveedor')
                almacen_id = request.POST.get('almacen')
                total = safe_decimal(request.POST.get('total', 0))
                fecha = request.POST.get('fecha') or timezone.now()
                observaciones = request.POST.get('observaciones', '')

                # Validar almacen
                almacen = Almacen.objects.get(
                    id=almacen_id,
                    sucursal_id=sucursal_id,
                    sucursal__fk_empresa=request.user.sucursal.fk_empresa
                )

                # Crear compra
                compra = Compra.objects.create(
                    usuario=usuario,
                    sucursal_id=sucursal_id,
                    proveedor_id=proveedor_id,
                    almacen=almacen,
                    total=total,
                    fecha=fecha,
                    observaciones=observaciones
                )

                # Detalles de productos
                productos = request.POST.getlist('producto[]')
                cantidades = request.POST.getlist('cantidad[]')
                precios = request.POST.getlist('precio[]')
                subtotales = request.POST.getlist('subtotal[]')

                for p_id, cant, precio, sub in zip(productos, cantidades, precios, subtotales):
                    if not p_id:
                        continue

                    cantidad = safe_decimal(cant)
                    precio_unitario = safe_decimal(precio)

                    # producto_variante_id, NO producto_id -> DetalleCompra.producto ya no existe
                    detalle = DetalleCompra.objects.create(
                        compra=compra,
                        producto_variante_id=p_id,
                        cantidad=cantidad,
                        precio=precio_unitario,
                        subtotal=safe_decimal(sub)
                    )

                    # 🔹 Kardex (mismo criterio que crear_venta: cantidad positiva,
                    #    tipo_movimiento como string, producto_variante_id no producto_id)
                    Kardex.objects.create(
                        producto_variante_id=p_id,
                        sucursal_id=sucursal_id,
                        almacen_id=almacen_id,
                        tipo_movimiento='entrada',
                        cantidad=cantidad,
                        precio_unitario=precio_unitario,
                        total=safe_decimal(sub),
                        referencia=f'Compra #{compra.id}'
                    )

                    # 🔹 ESTO FALTABA: crear_venta sí actualiza Stock, crear_compra
                    #    nunca lo hacía. Sin esto el inventario nunca sube con compras.
                    producto_variante = ProductoVariante.objects.get(id=p_id)
                    if producto_variante.maneja_stock:
                        stock, _ = Stock.objects.get_or_create(
                            almacen=almacen,
                            producto_variante=producto_variante,
                            defaults={
                                'cantidad_actual': 0,
                                'costo_unitario_promedio': 0,
                                'valor_total': 0,
                                'cajas_actual': 0,
                                'peso_neto_total': 0
                            }
                        )

                        # Recalcular Costo Promedio Ponderado (CPP).
                        # En la venta NO se toca el costo (correcto: la salida usa
                        # el costo que ya había). En la entrada SÍ hay que recalcularlo,
                        # si no, costo_unitario_promedio se queda en 0 para siempre.
                        valor_actual = stock.cantidad_actual * stock.costo_unitario_promedio
                        valor_entrante = cantidad * precio_unitario
                        nueva_cantidad = stock.cantidad_actual + cantidad

                        if nueva_cantidad > 0:
                            stock.costo_unitario_promedio = (valor_actual + valor_entrante) / nueva_cantidad

                        stock.cantidad_actual = nueva_cantidad
                        stock.valor_total = stock.cantidad_actual * stock.costo_unitario_promedio
                        stock.save()

                messages.success(request, f'✅ Compra #{compra.id} registrada correctamente.')
                return redirect(reverse('comprobante_compra', args=[compra.id]))

        except Exception as e:
            messages.error(request, f'❌ Error al registrar la compra: {e}')

    # GET → mostrar formulario
    producto = ProductoVariante.objects.filter(is_active=True).select_related('producto')
    productos = Producto.objects.filter(is_active=True, fk_empresa=request.user.sucursal.fk_empresa)
    sucursales = Sucursal.objects.filter(estado=True, fk_empresa=request.user.sucursal.fk_empresa)
    almacenes = []
    proveedores = Proveedor.objects.filter(is_active=True, empresa=request.user.sucursal.fk_empresa)

    return render(request, 'inventario/registro_compra.html', {
        'producto': producto,
        'productos': productos,
        'sucursales': sucursales,
        'almacenes': almacenes,
        'proveedores': proveedores,
        'fecha_actual': timezone.now().strftime('%Y-%m-%d')
    })
    
@login_required
@permiso_requerido('compra_list', 'eliminar')
def eliminar_compra(request):
    if request.method == 'POST':
        compra_id = request.POST.get('id')
        motivo = request.POST.get('motivo', '').strip()
        compra = get_object_or_404(Compra, id=compra_id)
        compra.is_active = False
        compra.motivo_anulacion = motivo
        compra.save()

        # Detalles inactivos
        compra.detallecompra_set.update(is_active=False)
        messages.success(request, f'Compra #{compra.id} eliminada correctamente.')
        return redirect('lista_compras')

@login_required
@permiso_requerido('compra_list', 'ver')
def comprobante_compra(request, compra_id):
    compra = get_object_or_404(Compra.objects.prefetch_related('detallecompra_set__producto_variante', 'sucursal', 'usuario', 'proveedor'), id=compra_id, is_active=True)
    empresa_nombre = request.user.sucursal.fk_empresa.nombre
    return render(request, 'inventario/comprobante_compra.html', {'compra': compra, 'empresa_nombre': empresa_nombre})

# ====================================================
#  VENTA (MAESTRO-DETALLE)
# ====================================================

@login_required
def ticket_cliente(request, venta_id):
    venta = get_object_or_404(Venta, pk=venta_id, is_active=True)
    sucursal = venta.sucursal
    empresa = sucursal.fk_empresa
    
    detalles = venta.detalles.filter(
        is_active=True, 
        detalle_padre__isnull=True
    ).select_related('producto_variante', 'producto_padre').prefetch_related('componentes')
    
    subtotal = sum(d.subtotal for d in detalles)
    
    for detalle in detalles:
        if detalle.producto_padre:
            detalle.componentes_list = detalle.componentes.filter(is_active=True)
    
    pagos = venta.pagos.select_related('metodo_pago').all()
    totalfinal = venta.total + venta.costo_envio
    total_pagado = sum(p.monto for p in pagos)
    
    context = {
        'venta': venta,
        'sucursal': sucursal,
        'empresa': empresa,
        'detalles': detalles,
        'subtotal': subtotal,
        'pagos': pagos,
        'total_pagado': total_pagado,
        'totalfinal': totalfinal
    }
    return render(request, 'ventas/ticket_cliente.html', context)
@login_required
def ticket_cocina(request, venta_id):
    """
    Ticket para cocina/producción:
    solo items + cantidades + canal + observaciones.
    Sin precios.
    """
    venta = get_object_or_404(Venta, pk=venta_id, is_active=True)
    sucursal = venta.sucursal

    detalles = venta.detalles.filter(is_active=True).select_related(
        'producto_variante', 'producto_padre', 'detalle_padre'
    )

    # Total de items (solo raíz, sin componentes de pack)
    total_items = detalles.filter(detalle_padre__isnull=True).count()

    context = {
        'venta':       venta,
        'sucursal':    sucursal,
        'detalles':    detalles,
        'total_items': total_items,
    }
    return render(request, 'ventas/ticket_cocina.html', context)

@login_required
@permiso_requerido('abrir_caja', 'ver')
def venta_list(request):
    empresa = request.user.fk_empresa
 
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    canal_id = request.GET.get('canal_id')
    sucursal_id = request.GET.get('sucursal_id')
    usuario_id = request.GET.get('usuario_id')
    estado_venta = request.GET.get('estado_venta')  # 'activa' | 'anulada' | '' (todas)
    cliente_q = request.GET.get('cliente', '').strip()
 
    ventas = (
        Venta.objects
        .filter(sucursal__fk_empresa=empresa)
        .select_related('usuario', 'sucursal', 'canal', 'cliente', 'caja_turno')
        .prefetch_related('pagos__metodo_pago')
        .order_by('-fecha')
    )
 
    # Filtro de fechas: si no mandan ninguna, por defecto SOLO HOY.
    if fecha_desde:
        try:
            ventas = ventas.filter(fecha__date__gte=datetime.strptime(fecha_desde, '%Y-%m-%d').date())
        except ValueError:
            pass
    if fecha_hasta:
        try:
            ventas = ventas.filter(fecha__date__lte=datetime.strptime(fecha_hasta, '%Y-%m-%d').date())
        except ValueError:
            pass
    if not fecha_desde and not fecha_hasta:
        ventas = ventas.filter(fecha__date=timezone.now().date())
 
    if canal_id and canal_id.isdigit():
        ventas = ventas.filter(canal_id=canal_id)
    if sucursal_id and sucursal_id.isdigit():
        ventas = ventas.filter(sucursal_id=sucursal_id)
    if usuario_id and usuario_id.isdigit():
        ventas = ventas.filter(usuario_id=usuario_id)
    if estado_venta == 'activa':
        ventas = ventas.filter(is_active=True)
    elif estado_venta == 'anulada':
        ventas = ventas.filter(is_active=False)
    if cliente_q:
        ventas = ventas.filter(
            Q(cliente__nombre__icontains=cliente_q) | Q(cliente__apellido__icontains=cliente_q)
        )
 
    resumen = ventas.filter(is_active=True).aggregate(
        total_monto=DjangoSum('total'),
        cantidad=Count('id'),
    )
 
    context = {
        'ventas': ventas,
        'canales': CanalVenta.objects.filter(fk_empresa=empresa, is_active=True),
        'sucursales': Sucursal.objects.filter(fk_empresa=empresa, estado=True),
        'usuarios': Usuario.objects.filter(sucursal__fk_empresa=empresa, is_active=True),
        'resumen': resumen,
        'fecha_desde': fecha_desde or '',
        'fecha_hasta': fecha_hasta or '',
        'canal_id': canal_id or '',
        'sucursal_id': sucursal_id or '',
        'usuario_id': usuario_id or '',
        'estado_venta': estado_venta or '',
        'cliente_q': cliente_q,
        'mostrando_hoy': not fecha_desde and not fecha_hasta,
        'titulo': 'Reporte de Ventas',
    }
    return render(request, 'ventas/venta_list.html', context)
 

def safe_decimal(value, default=Decimal('0.00')):
    try:
        if not value or value == '':
            return default
        return Decimal(str(value))
    except (InvalidOperation, TypeError):
        return default

def obtener_precio_producto(producto_variante, sucursal, canal):
    """Obtiene el precio del producto según: PrecioProducto > precio_referencial"""
    # Buscar precio configurado en PrecioProducto
    precio_config = producto_variante.precios.filter(
        sucursal=sucursal,
        canal=canal,
        activo=True
    ).order_by('-fecha').first()
    
    if precio_config:
        return precio_config.precio
    
    # Si no hay precio configurado, usar precio_referencial
    return producto_variante.precio_referencial

def validar_stock(producto_variante, almacen, cantidad):
    """Valida stock solo si el producto maneja stock"""
    if not producto_variante.maneja_stock:
        return True  # Si no maneja stock, siempre hay disponible
    
    try:
        stock = Stock.objects.get(
            almacen=almacen,
            producto_variante=producto_variante
        )
        return stock.cantidad_actual >= cantidad
    except Stock.DoesNotExist:
        return cantidad <= 0  # Si no existe registro y cantidad es 0, ok
@login_required
def obtener_stocks(request):
    almacen_id = request.GET.get('almacen')
    if not almacen_id:
        return JsonResponse([], safe=False)
    
    stocks = Stock.objects.filter(
        almacen_id=almacen_id,
        almacen__sucursal=request.user.sucursal
    )
    
    data = [{
        'producto_variante_id': s.producto_variante_id,
        'cantidad_actual': float(s.cantidad_actual)
    } for s in stocks]
    
    return JsonResponse(data, safe=False)

def actualizar_stock(producto_variante, almacen, cantidad, es_salida=True):
    """Actualiza stock solo si el producto maneja stock"""
    if not producto_variante.maneja_stock:
        return  # No hacer nada si no maneja stock
    
    stock, created = Stock.objects.get_or_create(
        almacen=almacen,
        producto_variante=producto_variante,
        defaults={
            'cantidad_actual': 0,
            'costo_unitario_promedio': 0,
            'valor_total': 0,
            'cajas_actual': 0,
            'peso_neto_total': 0
        }
    )
    
    if es_salida:
        stock.cantidad_actual -= cantidad
    else:
        stock.cantidad_actual += cantidad
    
    stock.save()

def api_precios_canal(request):
    canal_id = request.GET.get('canal')
    almacen_id = request.GET.get('almacen') # Capturamos el almacén desde el frontend
    
    if not canal_id:
        return JsonResponse({'ok': False, 'error': 'Falta el parámetro canal'}, status=400)
    
    try:
        # Filtros base obligatorios: canal y que el precio esté vigente
        filtros = {'canal_id': canal_id, 'activo': True}
        
        # Si vino el almacén, obtenemos su sucursal para afinar el precio geográfico
        if almacen_id:
            almacen = Almacen.objects.filter(id=almacen_id).first()
            if almacen:
                filtros['sucursal_id'] = almacen.sucursal_id

        # Consultamos la base de datos con los filtros
        precios_query = PrecioProducto.objects.filter(**filtros)
        
        precios_dict = {}
        for p in precios_query:
            # USAMOS producto_variante_id porque tu JS maneja IDs de variantes en las tarjetas
            precios_dict[p.producto_variante_id] = float(p.precio)
            
        return JsonResponse({
            'ok': True,
            'precios': precios_dict
        })
        
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)

@login_required
@permiso_requerido('abrir_caja', 'crear')
def crear_venta(request):
    usuario = request.user
    sucursal = usuario.sucursal

    # ========== PETICIÓN POST JSON ==========
    if request.method == 'POST':
        try:
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({'ok': False, 'error': 'Datos inválidos'}, status=400)

            with transaction.atomic():
                almacen_id      = data.get('almacen')
                canal_id        = data.get('canal')
                cliente_id      = data.get('cliente')
                total           = Decimal(str(data.get('total', 0)))
                descuento_total = Decimal(str(data.get('descuento', 0)))
                observaciones   = data.get('observaciones', '')
                costo_envio     = Decimal(str(data.get('costo_envio', 0)))
                direccion_entrega = data.get('direccion_entrega', '')
                telefono_entrega  = data.get('telefono_entrega', '')
                items           = data.get('items', [])
                pagos           = data.get('pagos', [])   # ← NUEVO: lista de pagos múltiples

                if not items:
                    return JsonResponse({'ok': False, 'error': 'No hay productos'}, status=400)

                # Validar que haya al menos un pago
                if not pagos:
                    return JsonResponse({'ok': False, 'error': 'Debe ingresar al menos un método de pago'}, status=400)

                # Validar que el total de pagos cubra (total - descuento + costo_envio)
                monto_a_cobrar = total 
                total_pagado   = sum(Decimal(str(p.get('monto', 0))) for p in pagos)
                if total_pagado < monto_a_cobrar:
                    return JsonResponse({
                        'ok': False,
                        'error': f'El monto pagado ({total_pagado}) es menor al total a cobrar ({monto_a_cobrar})'
                    }, status=400)

                # Obtener objetos
                almacen = Almacen.objects.get(id=almacen_id, sucursal=sucursal, is_active=True)
                canal   = CanalVenta.objects.get(id=canal_id, is_active=True, fk_empresa=sucursal.fk_empresa)

                cliente_obj = None
                if cliente_id and cliente_id != '':
                    try:
                        cliente_obj = Cliente.objects.get(id=cliente_id, fk_empresa=sucursal.fk_empresa)
                    except Cliente.DoesNotExist:
                        pass

                # Caja turno activa
                caja_turno = CajaTurno.objects.filter(
                    sucursal=sucursal,
                    usuario=usuario,
                    is_active=True,
                    fecha_cierre__isnull=True
                ).first()

                if not caja_turno:
                    return JsonResponse({'ok': False, 'error': 'No hay caja turno activa'}, status=400)

                # ========== CREAR VENTA ==========
                venta = Venta(
                    usuario=usuario,
                    sucursal=sucursal,
                    almacen=almacen,
                    canal=canal,
                    cliente=cliente_obj,
                    caja_turno=caja_turno,
                    total=total,
                    descuento=descuento_total,
                    costo_envio=costo_envio,
                    direccion_entrega=direccion_entrega,
                    telefono_entrega=telefono_entrega,
                    fecha=timezone.now(),
                    observaciones=observaciones,
                )
                venta.save()

                # ==========================
                # MOVIMIENTO DE CAJA: el total de la venta (sin separar envío aquí)
                # El egreso del delivery se registra aparte si el operador lo decide
                # ==========================
                MovimientoCaja.objects.create(
                    caja_turno=caja_turno,
                    tipo='VENTA',
                    monto=monto_a_cobrar,
                    referencia=str(venta.id),
                    descripcion=f'Venta #{venta.id}',
                    usuario=usuario
                )

                # ==========================
                # EGRESO DE CAJA: costo de envío (sale de caja para el delivery)
                # ==========================
                if costo_envio > 0:
                    MovimientoCaja.objects.create(
                        caja_turno=caja_turno,
                        tipo='EGRESO',
                        monto=costo_envio,
                        referencia=str(venta.id),
                        descripcion=f'Costo envío Venta #{venta.id}',
                        usuario=usuario
                    )

                # Procesar items
                for item in items:
                    tipo     = item.get('tipo', 'producto')
                    item_id  = item.get('id')
                    cantidad = Decimal(str(item.get('cantidad', 1)))
                    precio   = Decimal(str(item.get('precio', 0)))
                    descuento = Decimal(str(item.get('descuento', 0)))
                    subtotal = Decimal(str(item.get('subtotal', 0)))
                    nombre   = item.get('nombre', '')

                    if cantidad <= 0:
                        continue

                    if tipo == 'producto':
                        producto_variante = ProductoVariante.objects.get(id=item_id, is_active=True)

                        DetalleVenta.objects.create(
                            venta=venta,
                            producto_variante=producto_variante,
                            nombre_producto=nombre or producto_variante.nombre_variante,
                            cantidad=cantidad,
                            precio=precio,
                            subtotal=subtotal,
                            descuento=descuento
                        )

                        Kardex.objects.create(
                            producto_variante=producto_variante,
                            sucursal=sucursal,
                            almacen=almacen,
                            tipo_movimiento='salida',
                            cantidad=cantidad,
                            precio_unitario=precio,
                            total=subtotal,
                            referencia=f'Venta #{venta.id}'
                        )

                        if producto_variante.maneja_stock:
                            stock, _ = Stock.objects.get_or_create(
                                almacen=almacen,
                                producto_variante=producto_variante,
                                defaults={
                                    'cantidad_actual': 0,
                                    'costo_unitario_promedio': 0,
                                    'valor_total': 0,
                                    'cajas_actual': 0,
                                    'peso_neto_total': 0
                                }
                            )
                            stock.cantidad_actual -= cantidad
                            stock.save()

                    elif tipo == 'pack':
                        producto_padre = Producto.objects.get(id=item_id, is_active=True)
                        pack_variante  = ProductoVariante.objects.filter(producto=producto_padre, is_active=True).first()

                        DetalleVenta.objects.create(
                            venta=venta,
                            producto_variante=pack_variante,
                            producto_padre=producto_padre,
                            nombre_producto=nombre or producto_padre.nombre,
                            cantidad=cantidad,
                            precio=precio,
                            subtotal=subtotal,
                            descuento=descuento
                        )

                # ==========================
                # PAGOS MÚLTIPLES
                # ==========================
                for pago_data in pagos:
                    metodo_id  = pago_data.get('metodo_id')
                    monto_pago = Decimal(str(pago_data.get('monto', 0)))
                    referencia = pago_data.get('referencia', '')

                    if monto_pago <= 0:
                        continue

                    try:
                        metodo = MetodoPago.objects.get(id=metodo_id, empresa=sucursal.fk_empresa, estado=True)
                    except MetodoPago.DoesNotExist:
                        continue

                    PagoVenta.objects.create(
                        venta=venta,
                        metodo_pago=metodo,
                        monto=monto_pago,
                        referencia_pago=referencia
                    )

                return JsonResponse({
                    'ok': True,
                    'venta_id': venta.id,
                    'mensaje': f'✅ Venta #{venta.id} registrada correctamente'
                })

        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'ok': False, 'error': str(e)}, status=500)

    # ========== GET - Mostrar formulario ==========
    canal_qs      = CanalVenta.objects.filter(is_active=True, fk_empresa=sucursal.fk_empresa)
    categorias    = Category.objects.filter(is_active=True, fk_empresa=sucursal.fk_empresa)
    almacenes     = Almacen.objects.filter(sucursal=sucursal, is_active=True)
    clientes      = Cliente.objects.filter(estado=True, fk_empresa=sucursal.fk_empresa)
    metodos_pago  = MetodoPago.objects.filter(empresa=sucursal.fk_empresa, estado=True)
    tipos_ingreso = TipoIngreso.objects.filter(is_active=True, fk_empresa=sucursal.fk_empresa)
    tipos_egreso  = TipoEgreso.objects.filter(is_active=True, fk_empresa=sucursal.fk_empresa)

    variantes = ProductoVariante.objects.filter(
        is_active=True,
        producto__fk_empresa=sucursal.fk_empresa
    ).select_related('producto', 'producto__category')

    packs_ids = list(set(
        DetallePack.objects.filter(
            producto_padre__producto__fk_empresa=sucursal.fk_empresa
        ).values_list('producto_padre__producto_id', flat=True).distinct()
    ))
    packs = Producto.objects.filter(id__in=packs_ids, is_active=True) if packs_ids else Producto.objects.none()

    canal_default = canal_qs.first()

    precio_producto = {}
    for v in variantes:
        precio_producto[v.id] = float(obtener_precio_producto(v, sucursal, canal_default))

    precio_pack = {}
    for pack in packs:
        pack_variante = ProductoVariante.objects.filter(producto=pack, is_active=True).first()
        if pack_variante:
            precio_pack[pack.id] = float(obtener_precio_producto(pack_variante, sucursal, canal_default))
        else:
            precio_pack[pack.id] = 0.00

    context = {
        'usuario': usuario,
        'sucursal': sucursal,
        'canales': canal_qs,
        'almacenes': almacenes,
        'clientes': clientes,
        'metodos_pago': metodos_pago,
        'tipos_ingreso': tipos_ingreso,
        'tipos_egreso': tipos_egreso,
        'categorias': categorias,
        'productos': variantes,
        'packs': packs,
        'precio_producto': precio_producto,
        'precio_pack': precio_pack,
        'fecha_actual': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    return render(request, 'ventas/registro_venta.html', context)

@login_required
@permiso_requerido('cocina_kanban', 'ver')
def cocina_kanban(request):
    usuario = request.user
    sucursal = usuario.sucursal

    caja_turno = CajaTurno.objects.filter(
        sucursal=sucursal,
        is_active=True,
        fecha_cierre__isnull=True).first()

    ventas_registradas = []
    ventas_en_proceso = []
    ventas_despachadas = []

    if caja_turno:
        base_qs = Venta.objects.filter(
            caja_turno=caja_turno,
            is_active=True
        ).select_related(
            'canal', 'cliente', 'usuario'
        ).prefetch_related(
            'detalles'
        ).order_by('fecha')

        # ✅ Filtrar por estado_venta
        ventas_registradas = list(base_qs.filter(estado_venta='registrado'))
        ventas_en_proceso = list(base_qs.filter(estado_venta='en_proceso'))
        ventas_despachadas = list(base_qs.filter(estado_venta='despachado'))

    context = {
        'sucursal': sucursal,
        'caja_turno': caja_turno,
        'ventas_registradas': ventas_registradas,      # ← Cambié el nombre
        'ventas_en_proceso': ventas_en_proceso,
        'ventas_despachadas': ventas_despachadas,
    }
    return render(request, 'ventas/cocina_kanban.html', context)


@login_required
@permiso_requerido('cocina_kanban', 'editar')
def actualizar_estado_cocina(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body)
        venta_id = data.get('venta_id')
        nuevo_estado = data.get('nuevo_estado')

        # ✅ Estados válidos para el negocio
        ESTADOS_VALIDOS = ['registrado', 'en_proceso', 'despachado']
        if nuevo_estado not in ESTADOS_VALIDOS:
            return JsonResponse({'ok': False, 'error': 'Estado inválido'}, status=400)

        venta = Venta.objects.get(
            id=venta_id,
            sucursal=request.user.sucursal,
            is_active=True
        )
        
        # ✅ ACTUALIZAR estado_venta (el único campo)
        venta.estado_venta = nuevo_estado
        venta.save(update_fields=['estado_venta', 'updated_at'])

        return JsonResponse({
            'ok': True,
            'venta_id': venta.id,
            'estado': nuevo_estado
        })

    except Venta.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Venta no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@login_required
def kanban_pedidos_json(request):
    usuario = request.user
    sucursal = usuario.sucursal

    caja_turno = CajaTurno.objects.filter(
        sucursal=sucursal,
        is_active=True,
        fecha_cierre__isnull=True
    ).first()

    if not caja_turno:
        return JsonResponse({'ok': True, 'ventas': []})

    ventas = Venta.objects.filter(
        caja_turno=caja_turno,
        is_active=True
    ).select_related('canal', 'cliente', 'usuario').prefetch_related('detalles')

    data = []
    for v in ventas:
        detalles = []
        for d in v.detalles.filter(is_active=True, detalle_padre__isnull=True):
            detalles.append({
                'nombre': d.nombre_producto if d.nombre_producto else 'Producto',
                'cantidad': float(d.cantidad),
                'es_pack': bool(d.producto_padre),
            })
        
        data.append({
            'id': v.id,
            'estado_venta': v.estado_venta,  # ✅ Usar estado_venta
            'canal': v.canal.nombre if v.canal else 'Sin canal',
            'hora': v.fecha.strftime('%H:%M'),
            'cliente': v.cliente.nombre if v.cliente else 'General',
            'vendedor': v.usuario.nombre if v.usuario else 'Admin',
            'observaciones': v.observaciones or '',
            'detalles': detalles,
            'total_items': len(detalles),
        })

    return JsonResponse({'ok': True, 'ventas': data})

@csrf_exempt
def sse_nuevos_pedidos(request, sucursal_id):
    """Endpoint SSE para recibir notificaciones de nuevos pedidos"""
    
    def event_stream():
        ultimo_id = None
        
        while True:
            # Buscar pedidos nuevos (últimos 10 segundos)
            desde = time.time() - 10
            fecha_limite = datetime.now()
            
            nuevos = Venta.objects.filter(
                sucursal_id=sucursal_id,
                fecha__gte=datetime.now() - timedelta(seconds=10)
            ).exclude(
                estado_venta='despachado'
            ).order_by('-fecha')
            
            # Si hay pedidos nuevos y no es el mismo que el último enviado
            if nuevos.exists() and (ultimo_id != nuevos.first().id):
                ultimo_id = nuevos.first().id
                
                # Formatear datos del nuevo pedido
                data = {
                    'id': nuevos.first().id,
                    'estado': nuevos.first().estado_venta,
                    'timestamp': time.time()
                }
                yield f"data: {json.dumps(data)}\n\n"
            
            time.sleep(2)  # Revisar cada 2 segundos
    
    response = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'
    return response

@login_required
@permiso_requerido('abrir_caja', 'eliminar')
def venta_delete(request):
    """
    Anular venta = baja lógica + REVERSIÓN DE STOCK (esto faltaba) +
    registro en Kardex, igual que ya hicimos con compra_delete.
    Requiere motivo, ya no se anula sin explicar por qué.
    """
    if request.method == 'POST':
        try:
            with transaction.atomic():
                empresa = request.user.fk_empresa
                venta_id = request.POST.get('id')
                motivo = request.POST.get('motivo_anulacion', '').strip()
 
                venta = get_object_or_404(
                    Venta, pk=venta_id, sucursal__fk_empresa=empresa, is_active=True
                )
 
                if not motivo:
                    raise ValueError("Debes indicar el motivo de la anulación.")
 
                detalles = venta.detalles.filter(is_active=True, producto_variante__isnull=False)
 
                for detalle in detalles:
                    variante = detalle.producto_variante
                    if not variante.maneja_stock:
                        continue
 
                    stock, _ = Stock.objects.get_or_create(
                        almacen=venta.almacen,
                        producto_variante=variante,
                        defaults={'cantidad_actual': 0, 'costo_unitario_promedio': 0, 'valor_total': 0},
                    )
                    stock.cantidad_actual += detalle.cantidad
                    stock.valor_total = stock.cantidad_actual * stock.costo_unitario_promedio
                    stock.save()
 
                    Kardex.objects.create(
                        producto_variante=variante,
                        sucursal=venta.sucursal,
                        almacen=venta.almacen,
                        tipo_movimiento='anulacion_venta',
                        cantidad=detalle.cantidad,
                        precio_unitario=detalle.precio,
                        total=detalle.subtotal,
                        referencia=f'Anulación venta #{venta.id}',
                    )
 
                venta.is_active = False
                venta.motivo_anulacion = motivo
                venta.estado_venta = 'Anulada'
                venta.save()
                venta.detalles.update(is_active=False)
 
                messages.success(request, f'✅ Venta #{venta.id} anulada. Stock revertido correctamente.')
 
        except ValueError as e:
            messages.error(request, f'❌ {e}')
        except Exception as e:
            messages.error(request, f'❌ Error al anular la venta: {e}')
 
    return redirect('venta_list')


# ====================================================
#  TRASPASO (MAESTRO-DETALLE)
# ====================================================
@login_required
def traspaso_list(request):
    traspasos = Traspaso.objects.select_related('usuario', 'sucursal_origen', 'sucursal_destino').all().order_by('-created_at')
    return render(request, 'traspaso/list.html', {'traspasos': traspasos})

@login_required
@transaction.atomic
def traspaso_create(request):
    sucursales = Sucursal.objects.filter(estado=True)
    productos = Producto.objects.filter(is_active=True)
    usuarios = Usuario.objects.filter(is_active=True)
    if request.method == 'POST':
        usuario_id = request.POST.get('usuario') or request.user.id
        sucursal_origen_id = request.POST.get('sucursal_origen')
        sucursal_destino_id = request.POST.get('sucursal_destino')
        fecha = request.POST.get('fecha')
        observaciones = request.POST.get('observaciones') or ''

        traspaso = Traspaso.objects.create(
            usuario_id=usuario_id,
            sucursal_origen_id=sucursal_origen_id,
            sucursal_destino_id=sucursal_destino_id,
            fecha=fecha,
            observaciones=observaciones,
            total=0
        )

        productos_list = request.POST.getlist('producto')
        cantidades = request.POST.getlist('cantidad')
        precios = request.POST.getlist('precio')

        total = 0
        for i, prod_id in enumerate(productos_list):
            if not prod_id:
                continue
            cantidad = float(cantidades[i]) if i < len(cantidades) and cantidades[i] else 0
            precio = float(precios[i]) if i < len(precios) and precios[i] else 0
            subtotal = cantidad * precio
            DetalleTraspaso.objects.create(traspaso=traspaso, producto_id=prod_id, cantidad=cantidad, precio=precio, subtotal=subtotal)
            total += subtotal

            # restar del stock origen
            try:
                stock_or = Stock.objects.get(producto_id=prod_id, sucursal_id=sucursal_origen_id)
                stock_or.cantidad_actual = float(stock_or.cantidad_actual or 0) - cantidad
                stock_or.save()
            except Stock.DoesNotExist:
                pass

            # sumar al stock destino
            stock_dest, created = Stock.objects.get_or_create(producto_id=prod_id, sucursal_id=sucursal_destino_id, defaults={
                'cantidad_actual': cantidad, 'cajas_actual': 0, 'peso_neto_total': 0, 'costo_unitario_promedio': precio, 'valor_total': cantidad * precio
            })
            if not created:
                stock_dest.cantidad_actual = float(stock_dest.cantidad_actual or 0) + cantidad
                stock_dest.save()

        traspaso.total = total
        traspaso.save()
        messages.success(request, 'Traspaso registrado correctamente.')
        return redirect('traspaso_list')

    return render(request, 'traspaso/create.html', {'sucursales': sucursales, 'productos': productos, 'usuarios': usuarios})

@login_required
@transaction.atomic
def traspaso_edit(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        traspaso = get_object_or_404(Traspaso, pk=id)
        # editar campos maestro y reemplazar detalles (implementar reversión stock si querés)
        traspaso.sucursal_origen_id = request.POST.get('sucursal_origen')
        traspaso.sucursal_destino_id = request.POST.get('sucursal_destino')
        traspaso.fecha = request.POST.get('fecha')
        traspaso.observaciones = request.POST.get('observaciones') or ''
        traspaso.save()
        DetalleTraspaso.objects.filter(traspaso=traspaso).delete()
        productos_list = request.POST.getlist('producto')
        cantidades = request.POST.getlist('cantidad')
        precios = request.POST.getlist('precio')
        total = 0
        for i, prod_id in enumerate(productos_list):
            if not prod_id:
                continue
            cantidad = float(cantidades[i]) if i < len(cantidades) and cantidades[i] else 0
            precio = float(precios[i]) if i < len(precios) and precios[i] else 0
            subtotal = cantidad * precio
            DetalleTraspaso.objects.create(traspaso=traspaso, producto_id=prod_id, cantidad=cantidad, precio=precio, subtotal=subtotal)
            total += subtotal
        traspaso.total = total
        traspaso.save()
        messages.success(request, 'Traspaso actualizado correctamente.')
        return redirect('traspaso_list')
    else:
        id = request.GET.get('id')
        traspaso = get_object_or_404(Traspaso, pk=id)
        detalles = DetalleTraspaso.objects.filter(traspaso=traspaso)
        sucursales = Sucursal.objects.filter(estado=True)
        productos = Producto.objects.filter(is_active=True)
        return render(request, 'traspaso/edit.html', {'traspaso': traspaso, 'detalles': detalles, 'sucursales': sucursales, 'productos': productos})

@login_required
def traspaso_delete(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        traspaso = get_object_or_404(Traspaso, pk=id)
        traspaso.is_active = False
        traspaso.save()
        DetalleTraspaso.objects.filter(traspaso=traspaso).update(is_active=False)
        messages.success(request, 'Traspaso desactivado correctamente.')
    return redirect('traspaso_list')


def validar_anulacion(registro):
    """
    Devuelve (True, None) si se puede anular, o (False, "motivo") si no.
    Regla: no se puede anular si ya pasó el límite de horas configurado,
    tomando como referencia el cierre de la caja_turno (si existe y está cerrada)
    o la fecha de creación del propio registro.
    """
    if not registro.is_active:
        return False, "Este movimiento ya fue anulado anteriormente."

    limite_horas = getattr(settings, "HORAS_LIMITE_ANULACION_MOVIMIENTO", 24)

    # Punto de referencia: cierre de caja si existe, si no, la creación del registro
    caja = registro.caja_turno
    if caja is not None and getattr(caja, "fecha_cierre", None):
        referencia = caja.fecha_cierre
    else:
        referencia = registro.created_at

    limite = referencia + timedelta(hours=limite_horas)

    if timezone.now() > limite:
        return False, (
            f"No se puede anular: han pasado más de {limite_horas} horas "
            f"desde el registro/cierre de caja."
        )

    return True, None

# ====================================================
#  INGRESO MONETARIO
# ====================================================
@login_required
@permiso_requerido('reporte_ingresos', 'ver')
def reporte_ingresos(request):
    ingresos = IngresoMonetario.objects.select_related(
        "motivo", "usuario", "caja_turno"
    ).all().order_by("-fecha")

    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    motivo_id = request.GET.get("motivo")
    usuario_id = request.GET.get("usuario")
    caja_turno_id = request.GET.get("caja_turno")
    estado = request.GET.get("estado")
    q = request.GET.get("q")

    if fecha_inicio:
        ingresos = ingresos.filter(fecha__date__gte=parse_date(fecha_inicio))
    if fecha_fin:
        ingresos = ingresos.filter(fecha__date__lte=parse_date(fecha_fin))
    if motivo_id:
        ingresos = ingresos.filter(motivo_id=motivo_id)
    if usuario_id:
        ingresos = ingresos.filter(usuario_id=usuario_id)
    if caja_turno_id:
        ingresos = ingresos.filter(caja_turno_id=caja_turno_id)
    if estado == "activos":
        ingresos = ingresos.filter(is_active=True)
    elif estado == "anulados":
        ingresos = ingresos.filter(is_active=False)
    if q:
        ingresos = ingresos.filter(observaciones__icontains=q)

    total_general = ingresos.aggregate(total=DjangoSum("monto"))["total"] or 0
    total_activos = ingresos.filter(is_active=True).aggregate(
        total=DjangoSum("monto")
    )["total"] or 0
    total_anulados = ingresos.filter(is_active=False).aggregate(
        total=DjangoSum("monto")
    )["total"] or 0

    paginator = Paginator(ingresos, 25)
    page = paginator.get_page(request.GET.get("page"))
    tipos_ingreso = TipoIngreso.objects.filter(is_active=True, fk_empresa=request.user.sucursal.fk_empresa)
    contexto = {
        "ingresos": page,
        "total_general": total_general,
        "total_activos": total_activos,
        "total_anulados": total_anulados,
        "filtros": request.GET,
        "tipos_ingreso": tipos_ingreso,
    }
    return render(request, "empresa/reporte_ingresos.html", contexto)

@login_required
@permiso_requerido('reporte_ingresos', 'eliminar')
def anular_ingreso(request, pk):
    ingreso = get_object_or_404(IngresoMonetario, pk=pk)

    puede_anular, error = validar_anulacion(ingreso)
    if not puede_anular:
        messages.error(request, error)
        return redirect("reporte_ingresos")

    motivo_anulacion = request.POST.get("motivo_anulacion", "").strip()
    if not motivo_anulacion:
        messages.error(request, "Debes indicar el motivo de la anulación.")
        return redirect("reporte_ingresos")

    ingreso.is_active = False
    ingreso.motivo_anulacion = motivo_anulacion
    ingreso.save(update_fields=["is_active", "motivo_anulacion", "updated_at"])

    messages.success(request, "Ingreso anulado correctamente.")
    return redirect("reporte_ingresos")

@login_required
def crear_ingreso_monetario(request):

    if request.method != 'POST':
        return redirect('crear_venta')

    usuario = request.user

    caja_turno = CajaTurno.objects.filter(
        usuario=usuario,
        sucursal=usuario.sucursal,
        estado='ABIERTA',
        is_active=True
    ).first()

    if not caja_turno:
        messages.error(
            request,
            'No existe una caja abierta.'
        )
        return redirect('crear_venta')

    ingreso = IngresoMonetario.objects.create(
        fecha=timezone.now(),
        monto=request.POST.get('monto') or 0,
        motivo_id=request.POST.get('motivo'),
        observaciones=request.POST.get(
            'observaciones',
            ''
        ),
        usuario=usuario,
        caja_turno=caja_turno
    )

    MovimientoCaja.objects.create(
        caja_turno=caja_turno,
        tipo='INGRESO',
        monto=ingreso.monto,
        referencia=str(ingreso.id),
        descripcion=f'Ingreso monetario #{ingreso.id}',
        usuario=usuario
    )

    messages.success(
        request,
        'Ingreso monetario registrado.'
    )

    return redirect(
        'ticket_ingreso_monetario',
        ingreso.id
    )

@login_required
def ticket_ingreso_monetario(request, id):

    ingreso = get_object_or_404(
        IngresoMonetario,
        id=id
    )

    return render(
        request,
        'print/ingreso_ticket.html',
        {
            'titulo': 'Ingreso Monetario',
            'ingreso': ingreso
        }
    )
# ====================================================
#  EGRESO MONETARIO
# ====================================================
@login_required
@permiso_requerido('reporte_egresos', 'ver')
def reporte_egresos(request):
    egresos = EgresoMonetario.objects.select_related(
        "motivo", "usuario", "caja_turno"
    ).all().order_by("-fecha")

    # ---- Filtros ----
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    motivo_id = request.GET.get("motivo")
    usuario_id = request.GET.get("usuario")
    caja_turno_id = request.GET.get("caja_turno")
    estado = request.GET.get("estado")  # 'activos' / 'anulados' / '' (todos)
    q = request.GET.get("q")  # búsqueda libre en observaciones

    if fecha_inicio:
        egresos = egresos.filter(fecha__date__gte=parse_date(fecha_inicio))
    if fecha_fin:
        egresos = egresos.filter(fecha__date__lte=parse_date(fecha_fin))
    if motivo_id:
        egresos = egresos.filter(motivo_id=motivo_id)
    if usuario_id:
        egresos = egresos.filter(usuario_id=usuario_id)
    if caja_turno_id:
        egresos = egresos.filter(caja_turno_id=caja_turno_id)
    if estado == "activos":
        egresos = egresos.filter(is_active=True)
    elif estado == "anulados":
        egresos = egresos.filter(is_active=False)
    if q:
        egresos = egresos.filter(observaciones__icontains=q)

    # ---- Totales ----
    total_general = egresos.aggregate(total=DjangoSum("monto"))["total"] or 0
    total_activos = egresos.filter(is_active=True).aggregate(
        total=DjangoSum("monto")
    )["total"] or 0
    total_anulados = egresos.filter(is_active=False).aggregate(
        total=DjangoSum("monto")
    )["total"] or 0

    # ---- Paginación ----
    paginator = Paginator(egresos, 25)
    page = paginator.get_page(request.GET.get("page"))
    tipos_egreso = TipoEgreso.objects.filter(is_active=True, fk_empresa=request.user.sucursal.fk_empresa)
    contexto = {
        "egresos": page,
        "total_general": total_general,
        "total_activos": total_activos,
        "total_anulados": total_anulados,
        "filtros": request.GET,
        "tipos_egreso": tipos_egreso,
    }
    return render(request, "empresa/reporte_egresos.html", contexto)

@login_required
@permiso_requerido('reporte_egresos', 'eliminar')
def anular_egreso(request, pk):
    egreso = get_object_or_404(EgresoMonetario, pk=pk)

    puede_anular, error = validar_anulacion(egreso)
    if not puede_anular:
        messages.error(request, error)
        return redirect("reporte_egresos")

    motivo_anulacion = request.POST.get("motivo_anulacion", "").strip()
    if not motivo_anulacion:
        messages.error(request, "Debes indicar el motivo de la anulación.")
        return redirect("reporte_egresos")

    egreso.is_active = False
    egreso.motivo_anulacion = motivo_anulacion
    egreso.save(update_fields=["is_active", "motivo_anulacion", "updated_at"])

    messages.success(request, "Egreso anulado correctamente.")
    return redirect("reporte_egresos")

@login_required
def crear_egreso_monetario(request):

    if request.method != 'POST':
        return redirect('crear_venta')

    usuario = request.user

    caja_turno = CajaTurno.objects.filter(
        usuario=usuario,
        sucursal=usuario.sucursal,
        estado='ABIERTA',
        is_active=True
    ).first()

    if not caja_turno:
        messages.error(
            request,
            'No existe una caja abierta.'
        )
        return redirect('crear_venta')

    egreso = EgresoMonetario.objects.create(
        fecha=timezone.now(),
        monto=request.POST.get('monto') or 0,
        motivo_id=request.POST.get('motivo'),
        observaciones=request.POST.get(
            'observaciones',
            ''
        ),
        usuario=usuario,
        caja_turno=caja_turno
    )

    MovimientoCaja.objects.create(
        caja_turno=caja_turno,
        tipo='EGRESO',
        monto=egreso.monto,
        referencia=str(egreso.id),
        descripcion=f'Egreso monetario #{egreso.id}',
        usuario=usuario
    )

    messages.success(
        request,
        'Egreso monetario registrado.'
    )

    return redirect(
        'ticket_egreso_monetario',
        egreso.id
    )

@login_required
def ticket_egreso_monetario(request, id):

    egreso = get_object_or_404(
        EgresoMonetario,
        id=id
    )

    return render(
        request,
        'print/egreso_ticket.html',
        {
            'titulo': 'Egreso Monetario',
            'egreso': egreso
        }
    )
# ====================================================
#  PLANES (Membresías/Servicios)
# ====================================================
@login_required
def plan_list(request):
    empresa = request.user.sucursal.fk_empresa
    plans = Plan.objects.filter(fk_empresa=empresa, estado=True).order_by('-fecha_creacion')
    return render(request, 'membresia/lista_planes.html', {'plans': plans})

@login_required
def plan_create(request):
    
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        nombre = request.POST.get('nombre', '').strip()
        tarifa = request.POST.get('tarifa') or 0
        duracion = request.POST.get('membresia') # Viene del select del HTML
        descripcion = request.POST.get('descripcion', '').strip()

        if not nombre:
            messages.error(request, 'El nombre del plan es obligatorio.')
            return redirect('plan_list')

        if Plan.objects.filter(nombre__iexact=nombre, fk_empresa=empresa, estado=True).exists():
            messages.error(request, f'Ya tienes un plan llamado "{nombre}".')
            return redirect('plan_list')

        Plan.objects.create(
            nombre=nombre,
            tarifa=tarifa,
            duracion_dias=duracion, # Cambiado: antes decía membresia
            descripcion=descripcion,
            fk_empresa=empresa,
            # Agregamos los nuevos campos por defecto si quieres
            permite_congelar=request.POST.get('permite_congelar') == 'on',
            cantidad_dias_congelamiento=request.POST.get('cantidad_dias_congelamiento') or 0
        )
        messages.success(request, 'Plan creado correctamente.')
    return redirect('plan_list')

@login_required
def plan_edit(request):
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        id_plan = request.POST.get('id')
        p = get_object_or_404(Plan, pk=id_plan, fk_empresa=empresa)
        
        nombre = request.POST.get('nombre', '').strip()

        if Plan.objects.filter(nombre__iexact=nombre, fk_empresa=empresa, estado=True).exclude(id=p.id).exists():
            messages.error(request, f'Ya existe otro plan con el nombre "{nombre}".')
            return redirect('plan_list')

        p.nombre = nombre
        p.tarifa = request.POST.get('tarifa') or p.tarifa
        p.duracion_dias = request.POST.get('membresia') # Cambiado: antes decía membresia
        p.descripcion = request.POST.get('descripcion')
        p.permite_congelar = request.POST.get('permite_congelar') == 'on'
        p.cantidad_dias_congelamiento = request.POST.get('cantidad_dias_congelamiento') or 0
        p.save()
        
        messages.success(request, 'Plan actualizado correctamente.')
    return redirect('plan_list')

@login_required
def plan_delete(request):
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        id_plan = request.POST.get('id')
        p = get_object_or_404(Plan, pk=id_plan, fk_empresa=empresa)
        
        # OJO: Aquí ya no puedes filtrar por Cliente.fk_plan porque ese campo NO EXISTE
        # Por ahora lo hacemos simple, luego validaremos con la tabla Membresia
        p.estado = False 
        p.save()
        messages.success(request, 'Plan eliminado correctamente.')
        
    return redirect('plan_list')

#====================================================
#  CLIENTE
#====================================================
@login_required
def buscar_clientes(request):
    q = request.GET.get('q', '')
    if len(q) < 2:
        return JsonResponse({'ok': True, 'clientes': []})
    
    sucursal = request.user.sucursal
    clientes = Cliente.objects.filter(
        Q(fk_empresa=sucursal.fk_empresa) &
        Q(estado=True) &
        (Q(nombre__icontains=q) | 
         Q(apellido__icontains=q) | 
         Q(nro_documento__icontains=q) |
         Q(telefono__icontains=q) |
         Q(email__icontains=q))
    )[:15]
    
    data = [{
        'id': c.id,
        'nombre': c.nombre,
        'apellido': c.apellido,
        'nro_documento': c.nro_documento,
        'telefono': c.telefono,
        'email': c.email
    } for c in clientes]
    
    return JsonResponse({'ok': True, 'clientes': data})

@login_required
def crear_cliente(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        data = json.loads(request.body)
        sucursal = request.user.sucursal
        
        cliente = Cliente.objects.create(
            fk_empresa=sucursal.fk_empresa,
            nombre=data.get('nombre'),
            apellido=data.get('apellido'),
            nro_documento=data.get('documento', ''),
            telefono=data.get('telefono', ''),
            email=data.get('email', ''),
            estado=True
        )
        return JsonResponse({'ok': True, 'cliente_id': cliente.id})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)

@login_required
@permiso_requerido('cliente_list', 'ver')
def cliente_list(request):
    clientes = Cliente.objects.select_related('fk_empresa').filter(estado=True, fk_empresa=request.user.sucursal.fk_empresa).order_by('-fecha_creacion')
    return render(request, 'empresa/lista_clientes.html', {'clientes': clientes})

@login_required
@permiso_requerido('cliente_list', 'crear')
@transaction.atomic
def cliente_create(request):
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        nro_documento = request.POST.get('nro_documento', '').strip()
        nombre = request.POST.get('nombre', '').strip()

        # 1. Validaciones básicas de campos obligatorios
        if not nombre or not nro_documento:
            messages.error(request, 'El nombre y el número de documento son obligatorios.')
            return redirect('cliente_list')

        # 2. Validación de duplicados
        if Cliente.objects.filter(nro_documento=nro_documento, fk_empresa=empresa, estado=True).exists():
            messages.error(request, f'Ya existe un cliente con el documento {nro_documento}.')
            return redirect('cliente_list')

        # 3. Preparar el objeto (sin guardar en DB todavía)
        nuevo_cliente = Cliente(
            nombre=nombre,
            apellido=request.POST.get('apellido', '').strip(),
            nro_documento=nro_documento,
            fk_empresa=empresa,
            telefono=request.POST.get('telefono', '').strip(),
            email=request.POST.get('email', '').strip(),
            fecha_nacimiento=request.POST.get('fecha_nacimiento') or None,
        )

        # 4. MANEJO Y VALIDACIÓN DE LA FOTO
        if 'foto_perfil' in request.FILES:
            foto = request.FILES['foto_perfil']
            
            # Validar tamaño (2MB máximo)
            if foto.size > 2 * 1024 * 1024:
                messages.error(request, 'La foto es muy pesada. Máximo 2MB.')
                return redirect('cliente_list')
            
            nuevo_cliente.foto_perfil = foto

        # 5. Guardado final
        nuevo_cliente.save()
        messages.success(request, 'Cliente creado correctamente.')
        
    return redirect('cliente_list')

@login_required
@permiso_requerido('cliente_list', 'editar')
@transaction.atomic
def cliente_edit(request):
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        cliente_id = request.POST.get('id')
        
        c = get_object_or_404(Cliente, pk=cliente_id, fk_empresa=empresa)
        nro_documento = request.POST.get('nro_documento', '').strip()

        # 1. Validaciones de negocio primero
        if Cliente.objects.filter(nro_documento=nro_documento, fk_empresa=empresa, estado=True).exclude(id=c.id).exists():
            messages.error(request, f'Ya existe otro cliente con el documento {nro_documento}.')
            return redirect('cliente_list')

        # 2. Validación de la FOTO (Antes de asignar nada)
        if 'foto_perfil' in request.FILES:
            nueva_foto = request.FILES['foto_perfil']
            if nueva_foto.size > 2 * 1024 * 1024: # 2MB
                messages.error(request, 'La foto es muy pesada. El máximo permitido es 2MB.')
                return redirect('cliente_list')
            c.foto_perfil = nueva_foto # Solo se asigna si pasó el tamaño

        # 3. Asignación de los demás campos
        c.nombre = request.POST.get('nombre')
        c.apellido = request.POST.get('apellido')
        c.nro_documento = nro_documento
        c.telefono = request.POST.get('telefono')
        c.email = request.POST.get('email')
        c.fecha_nacimiento = request.POST.get('fecha_nacimiento') or None
        
        c.save()
        messages.success(request, 'Cliente actualizado correctamente.')
        
    return redirect('cliente_list')

@login_required
@transaction.atomic
@permiso_requerido('cliente_list', 'eliminar')
def cliente_delete(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        c = get_object_or_404(Cliente, pk=id, fk_empresa=request.user.sucursal.fk_empresa)
        c.estado = False
        c.save()
        messages.success(request, 'Cliente eliminado correctamente.')

    return redirect('cliente_list')

#====================================================
# Membresias
#====================================================77

@login_required
@permiso_requerido('membresias', 'ver')
def membresia_list(request):
    empresa = request.user.sucursal.fk_empresa
    # Traemos las membresías con select_related para que la tabla cargue rápido
    membresias = Membresia.objects.select_related('fk_cliente', 'fk_plan').filter(
        fk_empresa=empresa, 
        is_active=True
    ).order_by('-fecha_creacion')
    
    # Necesitamos clientes y planes para los selects de los modales
    clientes = Cliente.objects.filter(fk_empresa=empresa, estado=True)
    planes = Plan.objects.filter(fk_empresa=empresa, estado=True)
    
    context = {
        'membresias': membresias,
        'clientes': clientes,
        'planes': planes
    }
    return render(request, 'membresia/lista_membresias.html', context)

@login_required
@permiso_requerido('membresias', 'crear')
def membresia_create(request):
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        
        id_cliente = request.POST.get('fk_cliente')
        id_plan = request.POST.get('fk_plan')
        fecha_inicio = request.POST.get('fecha_inicio')
        estado = request.POST.get('pendiente', 'pendiente') # Por defecto activa si no viene nada

        cliente = get_object_or_404(Cliente, pk=id_cliente, fk_empresa=empresa)
        plan = get_object_or_404(Plan, pk=id_plan, fk_empresa=empresa)

        # Creamos la membresía (el modelo calculará la fecha_fin sola)
        Membresia.objects.create(
            fk_cliente=cliente,
            fk_plan=plan,
            fecha_inicio=fecha_inicio,
            estado=estado,
            fk_empresa=empresa
        )
        
        messages.success(request, f'Membresía asignada correctamente a {cliente.nombre}.')
    return redirect('membresia_list')

@login_required
@permiso_requerido('membresias', 'editar')
def membresia_edit(request):
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        id_membresia = request.POST.get('id')
        m = get_object_or_404(Membresia, pk=id_membresia, fk_empresa=empresa)
        
        # Actualizamos campos básicos
        m.estado = request.POST.get('estado')
        m.fecha_inicio = request.POST.get('fecha_inicio')
        
        # Si cambias el plan, forzamos que se recalcule la fecha_fin en el save
        nuevo_plan_id = request.POST.get('fk_plan')
        if int(nuevo_plan_id) != m.fk_plan.id:
            m.fk_plan = get_object_or_404(Plan, pk=nuevo_plan_id, fk_empresa=empresa)
            m.fecha_fin = None # Para que el save lo calcule de nuevo con el nuevo plan

        m.save()
        messages.success(request, 'Membresía actualizada correctamente.')
    return redirect('membresia_list')

@login_required
@permiso_requerido('membresias', 'eliminar')
def membresia_delete(request):
    
    if request.method == 'POST':
        empresa = request.user.sucursal.fk_empresa
        id_membresia = request.POST.get('id')
        m = get_object_or_404(Membresia, pk=id_membresia, fk_empresa=empresa)
        
        m.is_active = False # Soft delete
        m.save()
        messages.success(request, 'Membresía eliminada.')
    return redirect('membresia_list')

# ====================================================
#  PAGO Y DETALLE PAGO
# ====================================================
@login_required
def pago_list(request):
    pagos = Pago.objects.select_related('fk_cliente', 'fk_usuario').all().order_by('-fecha_creacion')
    return render(request, 'pago/list.html', {'pagos': pagos})

@login_required
@transaction.atomic
def pago_create(request):
    if request.method == 'POST':
        # 1. Datos del Maestro (Pago)
        cliente_id = request.POST.get('fk_cliente')
        metodo_id = request.POST.get('fk_metodo_pago')
        descuento = float(request.POST.get('descuento', 0))
        total = float(request.POST.get('total', 0))
        deuda = float(request.POST.get('deuda', 0))
        descripcion = request.POST.get('descripcion', '')
        
        # 2. Datos de los Detalles (Vienen como listas desde el form)
        planes_ids = request.POST.getlist('planes[]')
        cantidades = request.POST.getlist('cantidades[]')
        subtotales = request.POST.getlist('subtotales[]')

        # Creamos el objeto Pago
        # Nota: Como fk_membresia es obligatoria en tu modelo, 
        # primero deberías tener la membresía o ponerle null=True al modelo.
        # Por ahora asumo que se la asignamos a la membresía principal.
        pago = Pago.objects.create(
            nombre_pagador=request.POST.get('nombre_pagador'),
            fecha_hora_pago=timezone.now(),
            descripcion=descripcion,
            descuento=descuento,
            total=total,
            deuda=deuda,
            fk_metodo_pago_id=metodo_id,
            fk_usuario=request.user,
            fk_sucursal=request.user.sucursal,
            fk_membresia_id=request.POST.get('fk_membresia') # ID de la membresía que se está pagando
        )

        # 3. Registrar los detalles del pago
        for i in range(len(planes_ids)):
            DetallePago.objects.create(
                fk_pago=pago,
                fk_plan_id=planes_ids[i],
                cantidad=cantidades[i],
                subtotal=subtotales[i]
            )

        messages.success(request, f"Pago #{pago.id} registrado con éxito.")
        return redirect('pago_list')

    # Para cargar el formulario
    context = {
        'metodos': MetodoPago.objects.filter(empresa=request.user.sucursal.fk_empresa, estado=True),
        'planes': Plan.objects.filter(fk_empresa=request.user.sucursal.fk_empresa, estado=True),
    }
    return render(request, 'pagos/pago_form.html', context)

@login_required
def pago_edit(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        p = get_object_or_404(Pago, pk=id)
        p.fecha_hora_pago = request.POST.get('fecha_hora_pago')
        p.fecha_hora_inicio = request.POST.get('fecha_hora_inicio')
        p.descripcion = request.POST.get('descripcion')
        p.descuento = request.POST.get('descuento') or p.descuento
        p.total = request.POST.get('total') or p.total
        p.deuda = request.POST.get('deuda') or p.deuda
        p.fk_cliente_id = request.POST.get('fk_cliente') or p.fk_cliente_id
        p.fk_usuario_id = request.POST.get('fk_usuario') or p.fk_usuario_id
        p.save()
        messages.success(request, 'Pago actualizado correctamente.')
        return redirect('pago_list')
    else:
        id = request.GET.get('id')
        p = get_object_or_404(Pago, pk=id)
        clientes = Cliente.objects.filter(estado=True)
        usuarios = Usuario.objects.filter(is_active=True)
        return render(request, 'pago/edit.html', {'pago': p, 'clientes': clientes, 'usuarios': usuarios})

@login_required
def pago_delete(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        p = get_object_or_404(Pago, pk=id)
        p.estado = False
        p.save()
        messages.success(request, 'Pago desactivado correctamente.')
    return redirect('pago_list')


# ====================================================
#  ASISTENCIA
# ====================================================
@login_required
def asistencia_list(request):
    asistencias = Asistencia.objects.select_related('fk_cliente', 'fk_usuario', 'fk_sucursal').all().order_by('-fecha_creacion')
    return render(request, 'asistencia/list.html', {'asistencias': asistencias})

@login_required
def asistencia_create(request):
    clientes = Cliente.objects.filter(estado=True)
    usuarios = Usuario.objects.filter(is_active=True)
    sucursales = Sucursal.objects.filter(estado=True)
    if request.method == 'POST':
        fecha_hora = request.POST.get('fecha_hora')
        fk_cliente = request.POST.get('fk_cliente')
        fk_usuario = request.POST.get('fk_usuario') or None
        fk_sucursal = request.POST.get('fk_sucursal')
        Asistencia.objects.create(fecha_hora=fecha_hora, fk_cliente_id=fk_cliente, fk_usuario_id=fk_usuario, fk_sucursal_id=fk_sucursal)
        messages.success(request, 'Asistencia registrada correctamente.')
        return redirect('asistencia_list')
    return render(request, 'asistencia/create.html', {'clientes': clientes, 'usuarios': usuarios, 'sucursales': sucursales})

@login_required
def asistencia_edit(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        a = get_object_or_404(Asistencia, pk=id)
        a.fecha_hora = request.POST.get('fecha_hora') or a.fecha_hora
        a.fk_cliente_id = request.POST.get('fk_cliente') or a.fk_cliente_id
        a.fk_usuario_id = request.POST.get('fk_usuario') or a.fk_usuario_id
        a.fk_sucursal_id = request.POST.get('fk_sucursal') or a.fk_sucursal_id
        a.save()
        messages.success(request, 'Asistencia actualizada correctamente.')
        return redirect('asistencia_list')
    else:
        id = request.GET.get('id')
        a = get_object_or_404(Asistencia, pk=id)
        clientes = Cliente.objects.filter(estado=True)
        usuarios = Usuario.objects.filter(is_active=True)
        sucursales = Sucursal.objects.filter(estado=True)
        return render(request, 'asistencia/edit.html', {'asistencia': a, 'clientes': clientes, 'usuarios': usuarios, 'sucursales': sucursales})

@login_required
def asistencia_delete(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        a = get_object_or_404(Asistencia, pk=id)
        a.estado = False
        a.save()
        messages.success(request, 'Asistencia desactivada correctamente.')
    return redirect('asistencia_list')


# ====================================================
#  METODO DE PAGO
# ====================================================

@login_required
@permiso_requerido('metodos_pago', 'ver')
def metodopago_list(request):
    empresa_user = request.user.sucursal.fk_empresa
    # Solo vemos los métodos de MI empresa
    metodos = MetodoPago.objects.filter(
        empresa=empresa_user, 
        estado=True
    ).order_by('nombre')
    
    return render(request, 'membresia/lista_metodos.html', {'metodos': metodos})

@login_required
@permiso_requerido('metodos_pago', 'crear')
def metodopago_create(request):
    empresa = request.user.sucursal.fk_empresa
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()

        if not nombre:
            messages.error(request, 'El nombre del método de pago es obligatorio.')
            return redirect('metodopago_list')

        # VALIDACIÓN DE REPETIDOS (Ignora mayúsculas/minúsculas)
        if MetodoPago.objects.filter(nombre__iexact=nombre, estado=True, empresa=empresa).exists():
            messages.error(request, f'El método de pago "{nombre}" ya existe.')
            return redirect('metodopago_list')

        MetodoPago.objects.create(
            nombre=nombre,
            descripcion=descripcion,
            empresa=empresa 
        )
        messages.success(request, 'Método de pago creado correctamente.')
    return redirect('metodopago_list')

@login_required
@permiso_requerido('metodos_pago', 'editar')
def metodopago_edit(request):
    empresa_user = request.user.sucursal.fk_empresa
    if request.method == 'POST':
        id_metodo = request.POST.get('id')
        metodo_obj = get_object_or_404(MetodoPago, pk=id_metodo)
        
        nombre = request.POST.get('nombre', '').strip()
        
        # VALIDACIÓN DE REPETIDOS (Excluyendo al mismo que estamos editando)
        if MetodoPago.objects.filter(nombre__iexact=nombre, estado=True, empresa=empresa_user).exclude(id=metodo_obj.id).exists():
            messages.error(request, f'Ya existe otro método de pago con el nombre "{nombre}".')
            return redirect('metodopago_list')

        metodo_obj.nombre = nombre
        metodo_obj.descripcion = request.POST.get('descripcion', '').strip()
        metodo_obj.empresa = empresa_user
        metodo_obj.save()
        
        messages.success(request, 'Método de pago actualizado.')
    return redirect('metodopago_list')

@login_required
@permiso_requerido('metodos_pago', 'eliminar')
@transaction.atomic
def metodopago_delete(request):
    if request.method == 'POST':
        empresa_user = request.user.sucursal.fk_empresa

        id = request.POST.get('id')
        metodo = get_object_or_404(MetodoPago, pk=id, empresa=empresa_user)
        metodo.estado = False
        metodo.save()
        
        messages.success(request, 'Método de pago eliminado correctamente.')
    return redirect('metodopago_list')